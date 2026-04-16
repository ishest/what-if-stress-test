from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
import re
import time
from typing import Any, Callable

import math

import pandas as pd
import requests
import yfinance as yf
from openpyxl import load_workbook
from yahooquery import Ticker


WORKBOOK_PATH = Path(__file__).with_name("WhatIf_StressTest_v4_Fixed.xlsx")
MILLION = 1_000_000.0

HISTORICAL_LINE_ITEMS = [
    "Revenue",
    "COGS",
    "Gross Profit",
    "SG&A",
    "Other OpEx",
    "EBITDA",
    "D&A",
    "EBIT",
    "Interest Expense",
    "Pretax Income",
    "Taxes",
    "Net Income",
    "Cash & Equivalents",
    "Accounts Receivable",
    "Inventory",
    "Other Current Assets",
    "PP&E, net",
    "Intangibles & Goodwill",
    "Other Non-current Assets",
    "Short-term Debt",
    "Accounts Payable",
    "Other Current Liabilities",
    "Long-term Debt",
    "Other Non-current Liabilities",
    "Equity",
    "Balance Check",
    "Capex",
    "Dividends",
    "Buybacks",
    "CFO (actual, optional)",
    "Gross Margin %",
    "EBITDA Margin %",
    "EBIT Margin %",
    "Net Margin %",
    "Revenue YoY %",
    "EBITDA YoY %",
    "Net Leverage",
    "Interest Coverage",
    "DSO (days)",
    "DIO (days)",
    "DPO (days)",
    "FCF (CFO - Capex)",
    "Capex / Revenue %",
]

REQUIRED_BASE_FIELDS = [
    "Revenue",
    "COGS",
    "SG&A",
    "Other OpEx",
    "D&A",
    "Interest Expense",
    "Cash & Equivalents",
    "Accounts Receivable",
    "Accounts Payable",
    "Current Assets",
    "Current Liabilities",
    "PP&E, net",
    "Equity",
    "Capex",
]

STRESS_REQUIRED_FIELDS = [
    "Revenue",
    "COGS",
    "Gross Profit",
    "SG&A",
    "Other OpEx",
    "EBITDA",
    "D&A",
    "Interest Expense",
    "Cash & Equivalents",
    "Accounts Receivable",
    "Inventory",
    "PP&E, net",
    "Accounts Payable",
    "Equity",
    "Capex",
]

DISTRESS_MIN_ENDING_CASH = 0.0
DISTRESS_MIN_INTEREST_COVERAGE = 3.0
DISTRESS_MAX_NET_LEVERAGE = 3.0
DISTRESS_MIN_CURRENT_RATIO = 0.2
WATCH_MIN_INTEREST_COVERAGE = 5.0
WATCH_MAX_NET_LEVERAGE = 2.0
WATCH_MIN_CURRENT_RATIO = 1.0
CASH_PRESERVATION_MIN = 0.01
YAHOO_RETRY_DELAYS = (0.0, 0.4, 1.2)
YAHOO_HTTP_BASE = "https://query2.finance.yahoo.com"
YAHOO_HTTP_CRUMB_URL = f"{YAHOO_HTTP_BASE}/v1/test/getcrumb"
YAHOO_HTTP_CONSENT_URL = "https://fc.yahoo.com"
YAHOO_HTTP_TIMEOUT = 20
YAHOO_HTTP_USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
_LAST_GOOD_OVERVIEWS: dict[str, CompanyOverview] = {}


@dataclass
class ThresholdSettings:
    sga_variable_cost_share: float
    minimum_cash_buffer: float
    minimum_interest_coverage: float
    maximum_net_leverage: float
    minimum_current_ratio: float


@dataclass
class CompanyOverview:
    ticker: str
    short_name: str
    long_name: str
    sector: str
    industry: str
    exchange: str
    currency: str
    website: str
    market_cap_m: float | None
    current_price: float | None
    summary: str


@dataclass
class HistoricalDataset:
    overview: CompanyOverview
    annual_raw: pd.DataFrame
    financials: pd.DataFrame
    sources: pd.DataFrame
    latest_year: str
    latest_values: dict[str, float]
    blockers: list[str]
    warnings: list[str]
    data_quality_score: float
    sector_warning: str | None


class StressModelDataError(ValueError):
    def __init__(self, missing_fields: list[str], message: str | None = None):
        self.missing_fields = missing_fields
        detail = ", ".join(missing_fields)
        super().__init__(message or f"Stress model unavailable because required fields are missing: {detail}.")


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except Exception:
        return False


def _to_float(value: Any) -> float | None:
    if _is_missing(value):
        return None
    try:
        number = float(value)
    except Exception:
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def _to_millions(value: float | None, absolute: bool = False) -> float | None:
    if value is None:
        return None
    if absolute:
        value = abs(value)
    return value / MILLION


def _first_present(
    row: pd.Series,
    candidates: list[tuple[str, Callable[[pd.Series], float | None] | str]],
    *,
    absolute: bool = False,
    scale_to_millions: bool = True,
) -> tuple[float | None, str]:
    for label, candidate in candidates:
        raw_value = candidate(row) if callable(candidate) else _to_float(row.get(candidate))
        raw_value = _to_float(raw_value)
        if raw_value is None:
            continue
        value = _to_millions(raw_value, absolute=absolute) if scale_to_millions else abs(raw_value) if absolute else raw_value
        return value, label
    return None, "missing"


def _sum_raw(row: pd.Series, columns: list[str]) -> float | None:
    values = [_to_float(row.get(column)) for column in columns]
    clean_values = [value for value in values if value is not None]
    if not clean_values:
        return None
    return sum(clean_values)


def _subtract_raw(row: pd.Series, left: str, right: str) -> float | None:
    left_value = _to_float(row.get(left))
    right_value = _to_float(row.get(right))
    if left_value is None or right_value is None:
        return None
    return left_value - right_value


def _fallback_zero(label: str) -> tuple[float, str]:
    return 0.0, f"not reported -> 0 ({label})"


def _clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, value))


def _ratio(numerator: float | None, denominator: float | None, fallback: float = 0.0) -> float:
    if numerator is None or denominator is None or abs(denominator) < 1e-9:
        return fallback
    return numerator / denominator


def _symbol_payload(response: Any, symbol: str) -> dict[str, Any]:
    if not isinstance(response, dict):
        return {}
    payload = response.get(symbol, {})
    if isinstance(payload, dict):
        return payload
    return {}


def _yahoo_attempts():
    for delay in YAHOO_RETRY_DELAYS:
        if delay > 0:
            time.sleep(delay)
        yield


def _yahoo_raw(value: Any) -> Any:
    if isinstance(value, dict):
        raw = value.get("raw")
        if raw is not None:
            return raw
        if set(value.keys()) == {"fmt"}:
            return value.get("fmt")
    return value


def _empty_overview(symbol: str) -> CompanyOverview:
    return CompanyOverview(
        ticker=symbol,
        short_name=symbol,
        long_name=symbol,
        sector="",
        industry="",
        exchange="",
        currency="",
        website="",
        market_cap_m=None,
        current_price=None,
        summary="",
    )


def _create_yahoo_http_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": YAHOO_HTTP_USER_AGENT,
            "Accept": "application/json",
        }
    )
    return session


def _get_yahoo_http_crumb(session: requests.Session) -> str:
    try:
        session.get(YAHOO_HTTP_CONSENT_URL, allow_redirects=False, timeout=YAHOO_HTTP_TIMEOUT)
    except requests.RequestException:
        pass

    response = session.get(YAHOO_HTTP_CRUMB_URL, timeout=YAHOO_HTTP_TIMEOUT)
    response.raise_for_status()
    crumb = response.text.strip()
    if not crumb:
        raise ValueError("Yahoo Finance returned an empty crumb.")
    return crumb


def _fetch_quote_summary_modules(symbol: str, modules: list[str]) -> dict[str, Any]:
    last_error: Exception | None = None
    url = f"{YAHOO_HTTP_BASE}/v10/finance/quoteSummary/{symbol}"

    for attempt_index, delay in enumerate(YAHOO_RETRY_DELAYS):
        if delay > 0:
            time.sleep(delay)

        session = _create_yahoo_http_session()
        try:
            crumb = _get_yahoo_http_crumb(session)
            response = session.get(
                url,
                params={"modules": ",".join(modules), "crumb": crumb},
                timeout=YAHOO_HTTP_TIMEOUT,
            )
            if response.status_code == 429:
                raise requests.HTTPError("Yahoo Finance rate limited quoteSummary.", response=response)
            response.raise_for_status()
            payload = response.json()
            if not isinstance(payload, dict):
                raise ValueError("Yahoo Finance quoteSummary returned a non-dictionary payload.")

            quote_summary = payload.get("quoteSummary", {})
            if not isinstance(quote_summary, dict):
                raise ValueError("Yahoo Finance quoteSummary returned an invalid response envelope.")

            result = quote_summary.get("result") or []
            if result and isinstance(result[0], dict):
                return result[0]

            error = quote_summary.get("error")
            if error:
                if isinstance(error, dict):
                    description = error.get("description") or error.get("code") or str(error)
                else:
                    description = str(error)
                raise ValueError(f"Yahoo Finance quoteSummary returned no result for {symbol}: {description}")

            raise ValueError(f"Yahoo Finance quoteSummary returned no result for {symbol}.")
        except Exception as exc:
            last_error = exc
            if attempt_index == len(YAHOO_RETRY_DELAYS) - 1:
                break

    if last_error is not None:
        raise last_error
    raise ValueError(f"Yahoo Finance quoteSummary failed for {symbol}.")


def _fetch_quote_summary_modules_with_session(
    session: requests.Session,
    symbol: str,
    modules: list[str],
    crumb: str,
) -> dict[str, Any]:
    url = f"{YAHOO_HTTP_BASE}/v10/finance/quoteSummary/{symbol}"
    response = session.get(
        url,
        params={"modules": ",".join(modules), "crumb": crumb},
        timeout=YAHOO_HTTP_TIMEOUT,
    )
    if response.status_code == 429:
        raise requests.HTTPError("Yahoo Finance rate limited quoteSummary.", response=response)
    response.raise_for_status()
    payload = response.json()
    if not isinstance(payload, dict):
        raise ValueError("Yahoo Finance quoteSummary returned a non-dictionary payload.")

    quote_summary = payload.get("quoteSummary", {})
    if not isinstance(quote_summary, dict):
        raise ValueError("Yahoo Finance quoteSummary returned an invalid response envelope.")

    result = quote_summary.get("result") or []
    if result and isinstance(result[0], dict):
        return result[0]

    error = quote_summary.get("error")
    if error:
        if isinstance(error, dict):
            description = error.get("description") or error.get("code") or str(error)
        else:
            description = str(error)
        raise ValueError(f"Yahoo Finance quoteSummary returned no result for {symbol}: {description}")

    raise ValueError(f"Yahoo Finance quoteSummary returned no result for {symbol}.")


def _fetch_company_overview_via_http(symbol: str) -> CompanyOverview:
    quote_payload: dict[str, Any] = {}
    profile_payload: dict[str, Any] = {}
    last_error: Exception | None = None

    for attempt_index, delay in enumerate(YAHOO_RETRY_DELAYS):
        if delay > 0:
            time.sleep(delay)

        session = _create_yahoo_http_session()
        try:
            crumb = _get_yahoo_http_crumb(session)
        except Exception as exc:
            last_error = exc
            if attempt_index == len(YAHOO_RETRY_DELAYS) - 1:
                break
            continue

        try:
            quote_payload = _fetch_quote_summary_modules_with_session(session, symbol, ["price", "assetProfile"], crumb)
        except Exception as exc:
            last_error = exc
            quote_payload = {}

        try:
            profile_payload = _fetch_quote_summary_modules_with_session(
                session,
                symbol,
                ["financialData", "defaultKeyStatistics", "summaryDetail"],
                crumb,
            )
        except Exception as exc:
            last_error = exc
            profile_payload = {}

        if quote_payload or profile_payload:
            break

    if not quote_payload and not profile_payload:
        if last_error is not None:
            raise last_error
        raise ValueError(f"Yahoo Finance quoteSummary failed for {symbol}.")

    price = quote_payload.get("price", {}) if isinstance(quote_payload.get("price"), dict) else {}
    asset_profile = quote_payload.get("assetProfile", {}) if isinstance(quote_payload.get("assetProfile"), dict) else {}
    financial_data = profile_payload.get("financialData", {}) if isinstance(profile_payload.get("financialData"), dict) else {}

    market_cap = _to_float(_yahoo_raw(price.get("marketCap"))) or _to_float(_yahoo_raw(financial_data.get("marketCap")))
    current_price = _to_float(_yahoo_raw(price.get("regularMarketPrice"))) or _to_float(_yahoo_raw(financial_data.get("currentPrice")))

    short_name = str(price.get("shortName") or symbol)
    long_name = str(price.get("longName") or short_name or symbol)

    return CompanyOverview(
        ticker=symbol,
        short_name=short_name,
        long_name=long_name,
        sector=str(asset_profile.get("sector") or asset_profile.get("sectorDisp") or ""),
        industry=str(asset_profile.get("industry") or asset_profile.get("industryDisp") or ""),
        exchange=str(price.get("exchangeName") or price.get("fullExchangeName") or price.get("exchange") or ""),
        currency=str(price.get("currency") or financial_data.get("financialCurrency") or ""),
        website=str(asset_profile.get("website") or ""),
        market_cap_m=_to_millions(market_cap),
        current_price=current_price,
        summary=str(asset_profile.get("longBusinessSummary") or ""),
    )


def _fetch_company_overview_via_yahooquery(symbol: str) -> CompanyOverview:
    ticker = Ticker(symbol, asynchronous=False)

    try:
        quote_type_response = ticker.quote_type
    except Exception:
        quote_type_response = {}
    try:
        asset_profile_response = ticker.asset_profile
    except Exception:
        asset_profile_response = {}
    try:
        price_response = ticker.price
    except Exception:
        price_response = {}
    try:
        financial_data_response = ticker.financial_data
    except Exception:
        financial_data_response = {}

    quote_type = _symbol_payload(quote_type_response, symbol)
    asset_profile = _symbol_payload(asset_profile_response, symbol)
    price = _symbol_payload(price_response, symbol)
    financial_data = _symbol_payload(financial_data_response, symbol)

    market_cap = _to_float(price.get("marketCap")) or _to_float(financial_data.get("marketCap"))
    current_price = _to_float(price.get("regularMarketPrice")) or _to_float(financial_data.get("currentPrice"))
    short_name = str(quote_type.get("shortName") or price.get("shortName") or symbol)
    long_name = str(quote_type.get("longName") or price.get("longName") or short_name or symbol)

    return CompanyOverview(
        ticker=symbol,
        short_name=short_name,
        long_name=long_name,
        sector=str(asset_profile.get("sector") or ""),
        industry=str(asset_profile.get("industry") or ""),
        exchange=str(quote_type.get("exchange") or price.get("exchange") or ""),
        currency=str(price.get("currency") or financial_data.get("financialCurrency") or ""),
        website=str(asset_profile.get("website") or ""),
        market_cap_m=_to_millions(market_cap),
        current_price=current_price,
        summary=str(asset_profile.get("longBusinessSummary") or ""),
    )


def _fetch_company_overview_via_search(symbol: str) -> CompanyOverview:
    search = yf.Search(symbol, max_results=10)
    quotes = getattr(search, "quotes", None) or []
    quote = next(
        (item for item in quotes if str(item.get("symbol") or "").upper() == symbol.upper()),
        quotes[0] if quotes else {},
    )

    if not isinstance(quote, dict) or not quote:
        return _empty_overview(symbol)

    return CompanyOverview(
        ticker=symbol,
        short_name=str(quote.get("shortname") or symbol),
        long_name=str(quote.get("longname") or quote.get("shortname") or symbol),
        sector=str(quote.get("sectorDisp") or quote.get("sector") or ""),
        industry=str(quote.get("industryDisp") or quote.get("industry") or ""),
        exchange=str(quote.get("exchDisp") or quote.get("exchange") or ""),
        currency="",
        website="",
        market_cap_m=None,
        current_price=None,
        summary="",
    )


def _fetch_company_overview_via_yfinance(symbol: str) -> CompanyOverview:
    ticker = yf.Ticker(symbol)
    fast_info = getattr(ticker, "fast_info", None)
    metadata = ticker.get_history_metadata() or {}

    fast_price = None
    market_cap = None
    exchange = ""
    currency = ""
    if hasattr(fast_info, "get"):
        fast_price = _to_float(fast_info.get("lastPrice"))
        market_cap = _to_float(fast_info.get("marketCap"))
        exchange = str(fast_info.get("exchange") or "")
        currency = str(fast_info.get("currency") or "")

    current_price = fast_price or _to_float(metadata.get("regularMarketPrice")) or _to_float(metadata.get("previousClose"))
    short_name = str(metadata.get("shortName") or symbol)
    long_name = str(metadata.get("longName") or short_name or symbol)

    return CompanyOverview(
        ticker=symbol,
        short_name=short_name,
        long_name=long_name,
        sector="",
        industry="",
        exchange=exchange or str(metadata.get("exchangeName") or metadata.get("fullExchangeName") or ""),
        currency=currency or str(metadata.get("currency") or ""),
        website="",
        market_cap_m=_to_millions(market_cap),
        current_price=current_price,
        summary="",
    )


def _fetch_company_profile_enrichment(symbol: str) -> CompanyOverview:
    ticker = Ticker(symbol, asynchronous=False)

    try:
        asset_profile_response = ticker.asset_profile
    except Exception:
        asset_profile_response = {}

    asset_profile = _symbol_payload(asset_profile_response, symbol)
    return CompanyOverview(
        ticker=symbol,
        short_name=symbol,
        long_name=symbol,
        sector=str(asset_profile.get("sector") or ""),
        industry=str(asset_profile.get("industry") or ""),
        exchange="",
        currency="",
        website=str(asset_profile.get("website") or ""),
        market_cap_m=None,
        current_price=None,
        summary=str(asset_profile.get("longBusinessSummary") or ""),
    )


def _merge_company_overview(primary: CompanyOverview, secondary: CompanyOverview) -> CompanyOverview:
    symbol = primary.ticker or secondary.ticker

    def choose_text(primary_value: str, secondary_value: str, *, fallback_symbol: bool = False) -> str:
        cleaned_primary = (primary_value or "").strip()
        cleaned_secondary = (secondary_value or "").strip()
        if fallback_symbol and cleaned_primary == symbol and cleaned_secondary and cleaned_secondary != symbol:
            return cleaned_secondary
        return cleaned_primary or cleaned_secondary or (symbol if fallback_symbol else "")

    return CompanyOverview(
        ticker=symbol,
        short_name=choose_text(primary.short_name, secondary.short_name, fallback_symbol=True),
        long_name=choose_text(primary.long_name, secondary.long_name, fallback_symbol=True),
        sector=choose_text(primary.sector, secondary.sector),
        industry=choose_text(primary.industry, secondary.industry),
        exchange=choose_text(primary.exchange, secondary.exchange),
        currency=choose_text(primary.currency, secondary.currency),
        website=choose_text(primary.website, secondary.website),
        market_cap_m=primary.market_cap_m if primary.market_cap_m is not None else secondary.market_cap_m,
        current_price=primary.current_price if primary.current_price is not None else secondary.current_price,
        summary=choose_text(primary.summary, secondary.summary),
    )


def _overview_score(overview: CompanyOverview) -> int:
    score = 0
    if overview.current_price is not None:
        score += 1
    if overview.market_cap_m is not None:
        score += 1
    if bool((overview.summary or "").strip()):
        score += 1
    if overview.long_name and overview.long_name != overview.ticker:
        score += 1
    if overview.sector:
        score += 1
    if overview.industry:
        score += 1
    return score


def _remember_best_overview(symbol: str, overview: CompanyOverview) -> CompanyOverview:
    existing = _LAST_GOOD_OVERVIEWS.get(symbol)
    if existing is None:
        _LAST_GOOD_OVERVIEWS[symbol] = overview
        return overview

    merged = _merge_company_overview(overview, existing)
    best = merged if _overview_score(merged) >= _overview_score(existing) else existing
    _LAST_GOOD_OVERVIEWS[symbol] = best
    return best


def fetch_company_overview(symbol: str) -> CompanyOverview:
    symbol = symbol.upper().strip()
    try:
        overview = _fetch_company_overview_via_yfinance(symbol)
    except Exception:
        overview = None

    search_overview: CompanyOverview | None = None
    try:
        search_overview = _fetch_company_overview_via_search(symbol)
    except Exception:
        search_overview = None

    profile_overview: CompanyOverview | None = None
    try:
        profile_overview = _fetch_company_profile_enrichment(symbol)
    except Exception:
        profile_overview = None

    http_overview: CompanyOverview | None = None

    combined = overview or _empty_overview(symbol)
    if search_overview is not None:
        combined = _merge_company_overview(search_overview, combined)
    if profile_overview is not None:
        combined = _merge_company_overview(profile_overview, combined)

    # On some deployments the lightweight overview paths provide price/market cap
    # but not the long business summary. Use Yahoo's direct quoteSummary HTTP
    # endpoint only when the richer profile fields are still missing.
    if not (combined.summary or "").strip() or not combined.sector or not combined.industry or not combined.website:
        try:
            http_overview = _fetch_company_overview_via_http(symbol)
        except Exception:
            http_overview = None
        if http_overview is not None:
            combined = _merge_company_overview(http_overview, combined)

    if combined is not None:
        return _remember_best_overview(symbol, combined)
    if symbol in _LAST_GOOD_OVERVIEWS:
        return _LAST_GOOD_OVERVIEWS[symbol]
    return _empty_overview(symbol)


@lru_cache(maxsize=1)
def load_workbook_model(workbook_path: str = str(WORKBOOK_PATH)) -> tuple[pd.DataFrame, pd.DataFrame, ThresholdSettings]:
    scenario_library = pd.read_excel(workbook_path, sheet_name="Scenario_Library", header=2, engine="openpyxl")
    scenario_library = scenario_library.dropna(subset=["Sequence", "Severity"]).copy()
    scenario_library["Scenario Key"] = scenario_library["Scenario Key"].astype(str)

    sequence_map = pd.read_excel(workbook_path, sheet_name="WhatIf_Sequences", header=2, engine="openpyxl")
    sequence_map = sequence_map.dropna(subset=["Sequence"]).copy()

    workbook = load_workbook(workbook_path, data_only=False)
    setup_sheet = workbook["Scenario_Setup"]
    defaults = ThresholdSettings(
        sga_variable_cost_share=float(setup_sheet["B9"].value or 0.4),
        minimum_cash_buffer=float(setup_sheet["B11"].value or 0.0),
        minimum_interest_coverage=float(setup_sheet["B12"].value or 1.5),
        maximum_net_leverage=float(setup_sheet["B13"].value or 4.0),
        minimum_current_ratio=float(setup_sheet["B14"].value or 1.0),
    )
    return scenario_library, sequence_map, defaults


def _normalize_financial_label(label: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", str(label or ""))


def _coalesce_duplicate_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty or frame.columns.is_unique:
        return frame

    combined: dict[str, pd.Series] = {}
    for column in dict.fromkeys(map(str, frame.columns)):
        duplicate_frame = frame.loc[:, frame.columns == column]
        if isinstance(duplicate_frame, pd.Series):
            combined[column] = duplicate_frame
        else:
            combined[column] = duplicate_frame.bfill(axis=1).iloc[:, 0]
    return pd.DataFrame(combined, index=frame.index)


def _prepare_yfinance_statement_frame(statement: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(statement, pd.DataFrame) or statement.empty:
        return pd.DataFrame()

    frame = statement.T.copy()
    frame.index = pd.to_datetime(frame.index, errors="coerce")
    frame = frame[~frame.index.isna()]
    if frame.empty:
        return pd.DataFrame()

    frame.columns = [_normalize_financial_label(column) for column in frame.columns]
    frame = _coalesce_duplicate_columns(frame)
    frame.index.name = "asOfDate"
    return frame


def _fetch_annual_financials_via_yfinance(symbol: str) -> pd.DataFrame:
    ticker = yf.Ticker(symbol)
    statement_frames = []

    for attribute in ("income_stmt", "balance_sheet", "cash_flow"):
        try:
            prepared = _prepare_yfinance_statement_frame(getattr(ticker, attribute))
        except Exception:
            prepared = pd.DataFrame()
        if not prepared.empty:
            statement_frames.append(prepared)

    if not statement_frames:
        raise ValueError(f"No annual financial statement data returned by Yahoo Finance for ticker '{symbol}'.")

    combined = pd.concat(statement_frames, axis=1)
    combined = _coalesce_duplicate_columns(combined)
    annual = combined.reset_index().copy()
    if "asOfDate" not in annual.columns:
        raise ValueError(f"Yahoo Finance did not return annual statement dates for ticker '{symbol}'.")

    annual["asOfDate"] = pd.to_datetime(annual["asOfDate"], errors="coerce")
    annual = annual.dropna(subset=["asOfDate"]).sort_values("asOfDate")
    if annual.empty:
        raise ValueError(f"Yahoo Finance did not return usable annual statements for ticker '{symbol}'.")

    annual["Fiscal Year"] = annual["asOfDate"].dt.year.astype(str)
    latest_periods = annual.tail(min(12, len(annual))).copy()
    if latest_periods.shape[0] < 2:
        raise ValueError(f"Yahoo Finance returned fewer than two annual periods for ticker '{symbol}'.")
    return latest_periods.reset_index(drop=True)


def _fetch_annual_financials_via_yahooquery(symbol: str) -> pd.DataFrame:
    last_error: Exception | None = None

    for _ in _yahoo_attempts():
        try:
            ticker = Ticker(symbol, asynchronous=False)
            frame = ticker.all_financial_data()
            if not isinstance(frame, pd.DataFrame) or frame.empty:
                raise ValueError(f"No annual financial statement data returned by Yahoo Finance for ticker '{symbol}'.")

            annual = frame.reset_index().copy()
            if "asOfDate" not in annual.columns:
                raise ValueError(f"Yahoo Finance did not return annual statement dates for ticker '{symbol}'.")

            annual["asOfDate"] = pd.to_datetime(annual["asOfDate"], errors="coerce")
            annual = annual.dropna(subset=["asOfDate"])
            if "periodType" in annual.columns:
                annual = annual[annual["periodType"].astype(str).str.upper() == "12M"]
            annual = annual.sort_values("asOfDate")
            if annual.empty:
                raise ValueError(f"Yahoo Finance did not return usable 12M annual statements for ticker '{symbol}'.")

            annual["Fiscal Year"] = annual["asOfDate"].dt.year.astype(str)
            latest_periods = annual.tail(min(12, len(annual))).copy()
            if latest_periods.shape[0] < 2:
                raise ValueError(f"Yahoo Finance returned fewer than two annual periods for ticker '{symbol}'.")
            latest_periods = latest_periods.reset_index(drop=True)
            return latest_periods
        except Exception as exc:
            last_error = exc

    if last_error is not None:
        raise last_error
    raise ValueError(f"Yahoo Finance annual statement retrieval failed for ticker '{symbol}'.")


def fetch_annual_financials(symbol: str) -> pd.DataFrame:
    last_error: Exception | None = None

    for loader in (_fetch_annual_financials_via_yfinance, _fetch_annual_financials_via_yahooquery):
        try:
            return loader(symbol)
        except Exception as exc:
            last_error = exc

    if last_error is not None:
        raise last_error
    raise ValueError(f"Yahoo Finance annual statement retrieval failed for ticker '{symbol}'.")


def _map_single_year(row: pd.Series) -> tuple[dict[str, float | None], dict[str, str], list[str]]:
    warnings: list[str] = []
    values: dict[str, float | None] = {}
    sources: dict[str, str] = {}

    revenue, source = _first_present(row, [("TotalRevenue", "TotalRevenue"), ("OperatingRevenue", "OperatingRevenue")])

    gross_profit_direct, gp_source = _first_present(row, [("GrossProfit", "GrossProfit")])
    cogs, cogs_source = _first_present(
        row,
        [
            ("CostOfRevenue", "CostOfRevenue"),
            ("ReconciledCostOfRevenue", "ReconciledCostOfRevenue"),
            ("derived: Revenue - GrossProfit", lambda data: (_to_float(data.get("TotalRevenue")) or _to_float(data.get("OperatingRevenue"))) - _to_float(data.get("GrossProfit")) if (_to_float(data.get("TotalRevenue")) or _to_float(data.get("OperatingRevenue"))) is not None and _to_float(data.get("GrossProfit")) is not None else None),
        ],
        absolute=True,
    )
    values["COGS"], sources["COGS"] = cogs, cogs_source

    if revenue is None and gross_profit_direct is not None and cogs is not None:
        revenue = gross_profit_direct + cogs
        source = "derived: GrossProfit + COGS"
    values["Revenue"], sources["Revenue"] = revenue, source

    if revenue is not None and cogs is not None:
        values["Gross Profit"] = revenue - cogs
        sources["Gross Profit"] = "derived: Revenue - COGS"
    else:
        values["Gross Profit"] = gross_profit_direct
        sources["Gross Profit"] = gp_source

    sga, sga_source = _first_present(
        row,
        [
            ("SellingGeneralAndAdministration", "SellingGeneralAndAdministration"),
            ("derived: G&A + Sales & Marketing", lambda data: _sum_raw(data, ["GeneralAndAdministrativeExpense", "SellingAndMarketingExpense"])),
        ],
        absolute=True,
    )
    values["SG&A"], sources["SG&A"] = sga, sga_source

    reported_ebitda, reported_ebitda_source = _first_present(
        row,
        [("EBITDA", "EBITDA"), ("NormalizedEBITDA", "NormalizedEBITDA")],
    )
    operating_expense, _ = _first_present(
        row,
        [("OperatingExpense", "OperatingExpense")],
        absolute=True,
    )
    other_opex, other_opex_source = _first_present(
        row,
        [
            (
                "derived: OperatingExpense - SG&A",
                lambda data: _subtract_raw(data, "OperatingExpense", "SellingGeneralAndAdministration"),
            ),
            ("ResearchAndDevelopment", "ResearchAndDevelopment"),
            ("ResearchAndDevelopmentExpense", "ResearchAndDevelopmentExpense"),
            ("OtherOperatingExpenses", "OtherOperatingExpenses"),
            (
                "derived: GrossProfit - EBITDA - SG&A",
                lambda data: (_to_float(data.get("GrossProfit")) - _to_float(data.get("EBITDA")) - _to_float(data.get("SellingGeneralAndAdministration")))
                if _to_float(data.get("GrossProfit")) is not None
                and _to_float(data.get("EBITDA")) is not None
                and _to_float(data.get("SellingGeneralAndAdministration")) is not None
                else None,
            ),
        ],
        absolute=True,
    )
    if other_opex is None and operating_expense is not None and sga is None:
        other_opex = operating_expense
        other_opex_source = "OperatingExpense (SG&A not separately reported)"
        warnings.append("SG&A was not separately reported, so OperatingExpense is treated as Other OpEx.")
    values["Other OpEx"], sources["Other OpEx"] = other_opex, other_opex_source

    if values["Gross Profit"] is not None and sga is not None and other_opex is not None:
        values["EBITDA"] = values["Gross Profit"] - sga - other_opex
        sources["EBITDA"] = "derived: Gross Profit - SG&A - Other OpEx"
    else:
        values["EBITDA"] = reported_ebitda
        sources["EBITDA"] = reported_ebitda_source

    da, da_source = _first_present(
        row,
        [
            ("DepreciationAndAmortization", "DepreciationAndAmortization"),
            ("DepreciationAmortizationDepletion", "DepreciationAmortizationDepletion"),
            ("ReconciledDepreciation", "ReconciledDepreciation"),
            ("Depreciation", "Depreciation"),
        ],
        absolute=True,
    )
    values["D&A"], sources["D&A"] = da, da_source

    ebit_direct, ebit_direct_source = _first_present(row, [("EBIT", "EBIT")])
    if values["EBITDA"] is not None and da is not None:
        values["EBIT"] = values["EBITDA"] - da
        sources["EBIT"] = "derived: EBITDA - D&A"
    else:
        values["EBIT"] = ebit_direct
        sources["EBIT"] = ebit_direct_source

    interest_expense, interest_source = _first_present(
        row,
        [
            ("InterestExpense", "InterestExpense"),
            ("InterestExpenseNonOperating", "InterestExpenseNonOperating"),
            ("InterestPaidSupplementalData", "InterestPaidSupplementalData"),
        ],
        absolute=True,
    )
    values["Interest Expense"], sources["Interest Expense"] = interest_expense, interest_source

    pretax_direct, pretax_direct_source = _first_present(row, [("PretaxIncome", "PretaxIncome")])
    if values["EBIT"] is not None and interest_expense is not None:
        values["Pretax Income"] = values["EBIT"] - interest_expense
        sources["Pretax Income"] = "derived: EBIT - Interest Expense"
    else:
        values["Pretax Income"] = pretax_direct
        sources["Pretax Income"] = pretax_direct_source

    taxes, taxes_source = _first_present(
        row,
        [
            ("TaxProvision", "TaxProvision"),
            (
                "derived: PretaxIncome * TaxRateForCalcs",
                lambda data: _to_float(data.get("PretaxIncome")) * _to_float(data.get("TaxRateForCalcs"))
                if _to_float(data.get("PretaxIncome")) is not None and _to_float(data.get("TaxRateForCalcs")) is not None
                else None,
            ),
        ],
        absolute=True,
    )
    values["Taxes"], sources["Taxes"] = taxes, taxes_source

    net_income, net_income_source = _first_present(
        row,
        [
            ("NetIncome", "NetIncome"),
            ("NetIncomeCommonStockholders", "NetIncomeCommonStockholders"),
            ("NetIncomeContinuousOperations", "NetIncomeContinuousOperations"),
        ],
    )
    if net_income is None and values["Pretax Income"] is not None and taxes is not None:
        net_income = values["Pretax Income"] - taxes
        net_income_source = "derived: Pretax Income - Taxes"
    values["Net Income"], sources["Net Income"] = net_income, net_income_source

    cash, cash_source = _first_present(
        row,
        [
            ("CashAndCashEquivalents", "CashAndCashEquivalents"),
            ("CashCashEquivalentsAndShortTermInvestments", "CashCashEquivalentsAndShortTermInvestments"),
            ("EndCashPosition", "EndCashPosition"),
        ],
    )
    values["Cash & Equivalents"], sources["Cash & Equivalents"] = cash, cash_source

    receivables, receivables_source = _first_present(
        row,
        [
            ("AccountsReceivable", "AccountsReceivable"),
            (
                "derived: GrossAccountsReceivable - Allowance",
                lambda data: _to_float(data.get("GrossAccountsReceivable")) - abs(_to_float(data.get("AllowanceForDoubtfulAccountsReceivable") or 0))
                if _to_float(data.get("GrossAccountsReceivable")) is not None
                else None,
            ),
        ],
    )
    values["Accounts Receivable"], sources["Accounts Receivable"] = receivables, receivables_source

    inventory, inventory_source = _first_present(
        row,
        [
            ("Inventory", "Inventory"),
            ("derived: FinishedGoods + RawMaterials + OtherInventories", lambda data: _sum_raw(data, ["FinishedGoods", "RawMaterials", "OtherInventories"])),
        ],
    )
    if inventory is None:
        inventory, inventory_source = _fallback_zero("Inventory")
        warnings.append("Inventory was not reported and is treated as zero.")
    values["Inventory"], sources["Inventory"] = inventory, inventory_source

    current_assets, current_assets_source = _first_present(row, [("CurrentAssets", "CurrentAssets")])
    values["Current Assets"], sources["Current Assets"] = current_assets, current_assets_source

    ppe, ppe_source = _first_present(row, [("NetPPE", "NetPPE"), ("Properties", "Properties")])
    values["PP&E, net"], sources["PP&E, net"] = ppe, ppe_source

    intangibles, intangibles_source = _first_present(
        row,
        [
            ("GoodwillAndOtherIntangibleAssets", "GoodwillAndOtherIntangibleAssets"),
            ("derived: Goodwill + OtherIntangibleAssets", lambda data: _sum_raw(data, ["Goodwill", "OtherIntangibleAssets"])),
        ],
    )
    if intangibles is None:
        intangibles, intangibles_source = _fallback_zero("Intangibles & Goodwill")
    values["Intangibles & Goodwill"], sources["Intangibles & Goodwill"] = intangibles, intangibles_source

    total_assets, total_assets_source = _first_present(row, [("TotalAssets", "TotalAssets")])
    values["Total Assets"], sources["Total Assets"] = total_assets, total_assets_source

    st_debt, st_debt_source = _first_present(
        row,
        [
            ("CurrentDebt", "CurrentDebt"),
            ("CurrentDebtAndCapitalLeaseObligation", "CurrentDebtAndCapitalLeaseObligation"),
            ("OtherCurrentBorrowings", "OtherCurrentBorrowings"),
        ],
    )
    if st_debt is None:
        st_debt, st_debt_source = _fallback_zero("Short-term Debt")
    values["Short-term Debt"], sources["Short-term Debt"] = st_debt, st_debt_source

    payables, payables_source = _first_present(row, [("AccountsPayable", "AccountsPayable"), ("Payables", "Payables")])
    values["Accounts Payable"], sources["Accounts Payable"] = payables, payables_source

    current_liabilities, current_liabilities_source = _first_present(row, [("CurrentLiabilities", "CurrentLiabilities")])
    values["Current Liabilities"], sources["Current Liabilities"] = current_liabilities, current_liabilities_source

    lt_debt, lt_debt_source = _first_present(
        row,
        [("LongTermDebt", "LongTermDebt"), ("LongTermDebtAndCapitalLeaseObligation", "LongTermDebtAndCapitalLeaseObligation")],
    )
    if lt_debt is None:
        lt_debt, lt_debt_source = _fallback_zero("Long-term Debt")
    values["Long-term Debt"], sources["Long-term Debt"] = lt_debt, lt_debt_source

    total_liabilities, total_liabilities_source = _first_present(
        row,
        [("TotalLiabilitiesNetMinorityInterest", "TotalLiabilitiesNetMinorityInterest"), ("TotalLiabilities", "TotalLiabilities")],
    )
    values["Total Liabilities"], sources["Total Liabilities"] = total_liabilities, total_liabilities_source

    equity, equity_source = _first_present(
        row,
        [
            ("CommonStockEquity", "CommonStockEquity"),
            ("StockholdersEquity", "StockholdersEquity"),
            ("TotalEquityGrossMinorityInterest", "TotalEquityGrossMinorityInterest"),
        ],
    )
    values["Equity"], sources["Equity"] = equity, equity_source

    if current_assets is not None and cash is not None and receivables is not None and inventory is not None:
        values["Other Current Assets"] = current_assets - cash - receivables - inventory
        sources["Other Current Assets"] = "derived: CurrentAssets - Cash - A/R - Inventory"
    else:
        values["Other Current Assets"] = None
        sources["Other Current Assets"] = "missing"

    if total_assets is not None and current_assets is not None and ppe is not None and intangibles is not None:
        values["Other Non-current Assets"] = total_assets - current_assets - ppe - intangibles
        sources["Other Non-current Assets"] = "derived: TotalAssets - CurrentAssets - PP&E - Intangibles"
    else:
        values["Other Non-current Assets"] = None
        sources["Other Non-current Assets"] = "missing"

    if current_liabilities is not None and payables is not None and st_debt is not None:
        values["Other Current Liabilities"] = current_liabilities - payables - st_debt
        sources["Other Current Liabilities"] = "derived: CurrentLiabilities - A/P - Short-term Debt"
    else:
        values["Other Current Liabilities"] = None
        sources["Other Current Liabilities"] = "missing"

    if total_liabilities is not None and current_liabilities is not None and lt_debt is not None:
        values["Other Non-current Liabilities"] = total_liabilities - current_liabilities - lt_debt
        sources["Other Non-current Liabilities"] = "derived: TotalLiabilities - CurrentLiabilities - Long-term Debt"
    else:
        values["Other Non-current Liabilities"] = None
        sources["Other Non-current Liabilities"] = "missing"

    capex, capex_source = _first_present(
        row,
        [
            ("CapitalExpenditure", "CapitalExpenditure"),
            ("CapitalExpenditureReported", "CapitalExpenditureReported"),
            ("PurchaseOfPPE", "PurchaseOfPPE"),
        ],
        absolute=True,
    )
    values["Capex"], sources["Capex"] = capex, capex_source

    dividends, dividends_source = _first_present(
        row,
        [("CommonStockDividendPaid", "CommonStockDividendPaid"), ("CashDividendsPaid", "CashDividendsPaid")],
        absolute=True,
    )
    if dividends is None:
        dividends, dividends_source = _fallback_zero("Dividends")
    values["Dividends"], sources["Dividends"] = dividends, dividends_source

    buybacks, buybacks_source = _first_present(
        row,
        [("RepurchaseOfCapitalStock", "RepurchaseOfCapitalStock"), ("CommonStockPayments", "CommonStockPayments")],
        absolute=True,
    )
    if buybacks is None:
        buybacks, buybacks_source = _fallback_zero("Buybacks")
    values["Buybacks"], sources["Buybacks"] = buybacks, buybacks_source

    cfo, cfo_source = _first_present(
        row,
        [("OperatingCashFlow", "OperatingCashFlow"), ("CashFlowFromContinuingOperatingActivities", "CashFlowFromContinuingOperatingActivities")],
    )
    values["CFO (actual, optional)"], sources["CFO (actual, optional)"] = cfo, cfo_source

    values["Balance Check"] = None
    sources["Balance Check"] = "formula"
    return values, sources, warnings


def map_historical_financials(annual_data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    year_values: dict[str, dict[str, float | None]] = {}
    year_sources: dict[str, dict[str, str]] = {}
    warnings: list[str] = []

    for _, row in annual_data.iterrows():
        fiscal_year = str(row["Fiscal Year"])
        mapped_values, mapped_sources, row_warnings = _map_single_year(row)
        warnings.extend([f"{fiscal_year}: {warning}" for warning in row_warnings])
        year_values[fiscal_year] = mapped_values
        year_sources[fiscal_year] = mapped_sources

    financials = pd.DataFrame(year_values).reindex(HISTORICAL_LINE_ITEMS)
    sources = pd.DataFrame(year_sources).reindex(HISTORICAL_LINE_ITEMS)

    years = list(financials.columns)
    for idx, year in enumerate(years):
        revenue = financials.at["Revenue", year]
        cogs = financials.at["COGS", year]
        gross_profit = financials.at["Gross Profit", year]
        sga = financials.at["SG&A", year]
        other_opex = financials.at["Other OpEx", year]
        ebitda = financials.at["EBITDA", year]
        da = financials.at["D&A", year]
        ebit = financials.at["EBIT", year]
        interest = financials.at["Interest Expense", year]
        pretax = financials.at["Pretax Income", year]
        taxes = financials.at["Taxes", year]
        net_income = financials.at["Net Income", year]
        cash = financials.at["Cash & Equivalents", year]
        receivables = financials.at["Accounts Receivable", year]
        inventory = financials.at["Inventory", year]
        oca = financials.at["Other Current Assets", year]
        ppe = financials.at["PP&E, net", year]
        intangibles = financials.at["Intangibles & Goodwill", year]
        onca = financials.at["Other Non-current Assets", year]
        st_debt = financials.at["Short-term Debt", year]
        ap = financials.at["Accounts Payable", year]
        ocl = financials.at["Other Current Liabilities", year]
        lt_debt = financials.at["Long-term Debt", year]
        oncl = financials.at["Other Non-current Liabilities", year]
        equity = financials.at["Equity", year]
        capex = financials.at["Capex", year]
        dividends = financials.at["Dividends", year]
        buybacks = financials.at["Buybacks", year]
        cfo = financials.at["CFO (actual, optional)", year]

        if capex is None and idx > 0:
            prev_year = years[idx - 1]
            prev_ppe = financials.at["PP&E, net", prev_year]
            if ppe is not None and prev_ppe is not None and da is not None:
                derived_capex = max(0.0, ppe - prev_ppe + da)
                capex = derived_capex
                financials.at["Capex", year] = capex
                sources.at["Capex", year] = "derived approx: ΔPP&E + D&A"
                warnings.append(
                    f"{year}: Capex missing in Yahoo annuals; approximated from PP&E roll-forward (ΔPP&E + D&A)."
                )

        if gross_profit is None and revenue is not None and cogs is not None:
            gross_profit = revenue - cogs
            financials.at["Gross Profit", year] = gross_profit
            sources.at["Gross Profit", year] = "derived: Revenue - COGS"

        if ebitda is None and gross_profit is not None and sga is not None and other_opex is not None:
            ebitda = gross_profit - sga - other_opex
            financials.at["EBITDA", year] = ebitda
            sources.at["EBITDA", year] = "derived: Gross Profit - SG&A - Other OpEx"

        if ebit is None and ebitda is not None and da is not None:
            ebit = ebitda - da
            financials.at["EBIT", year] = ebit
            sources.at["EBIT", year] = "derived: EBITDA - D&A"

        if pretax is None and ebit is not None and interest is not None:
            pretax = ebit - interest
            financials.at["Pretax Income", year] = pretax
            sources.at["Pretax Income", year] = "derived: EBIT - Interest Expense"

        if net_income is None and pretax is not None and taxes is not None:
            net_income = pretax - taxes
            financials.at["Net Income", year] = net_income
            sources.at["Net Income", year] = "derived: Pretax Income - Taxes"

        if oca is None and None not in (cash, receivables, inventory):
            oca = 0.0
            financials.at["Other Current Assets", year] = oca
            sources.at["Other Current Assets", year] = "missing -> 0"

        if ocl is None and None not in (ap, st_debt):
            ocl = 0.0
            financials.at["Other Current Liabilities", year] = ocl
            sources.at["Other Current Liabilities", year] = "missing -> 0"

        if onca is None and None not in (ppe, intangibles):
            onca = 0.0
            financials.at["Other Non-current Assets", year] = onca
            sources.at["Other Non-current Assets", year] = "missing -> 0"

        if oncl is None and lt_debt is not None:
            oncl = 0.0
            financials.at["Other Non-current Liabilities", year] = oncl
            sources.at["Other Non-current Liabilities", year] = "missing -> 0"

        if None not in (cash, receivables, inventory, oca, ppe, intangibles, onca, st_debt, ap, ocl, lt_debt, oncl, equity):
            balance_check = (cash + receivables + inventory + oca + ppe + intangibles + onca) - (st_debt + ap + ocl + lt_debt + oncl + equity)
            financials.at["Balance Check", year] = balance_check
            sources.at["Balance Check", year] = "derived"

        gross_margin = _ratio(gross_profit, revenue)
        ebitda_margin = _ratio(ebitda, revenue)
        ebit_margin = _ratio(ebit, revenue)
        net_margin = _ratio(net_income, revenue)
        net_leverage = _ratio((st_debt or 0) + (lt_debt or 0) - (cash or 0), ebitda, fallback=0.0)
        interest_coverage = _ratio(ebit, interest, fallback=999.0)
        dso = _ratio((receivables or 0) * 365, revenue, fallback=0.0)
        dio = _ratio((inventory or 0) * 365, cogs, fallback=0.0)
        dpo = _ratio((ap or 0) * 365, cogs, fallback=0.0)
        fcf = (cfo if cfo is not None else (net_income or 0) + (da or 0)) - (capex or 0)
        capex_pct = _ratio(capex, revenue)

        financials.at["Gross Margin %", year] = gross_margin
        financials.at["EBITDA Margin %", year] = ebitda_margin
        financials.at["EBIT Margin %", year] = ebit_margin
        financials.at["Net Margin %", year] = net_margin
        financials.at["Net Leverage", year] = net_leverage
        financials.at["Interest Coverage", year] = interest_coverage
        financials.at["DSO (days)", year] = dso
        financials.at["DIO (days)", year] = dio
        financials.at["DPO (days)", year] = dpo
        financials.at["FCF (CFO - Capex)", year] = fcf
        financials.at["Capex / Revenue %", year] = capex_pct

        for ratio_label in [
            "Gross Margin %",
            "EBITDA Margin %",
            "EBIT Margin %",
            "Net Margin %",
            "Net Leverage",
            "Interest Coverage",
            "DSO (days)",
            "DIO (days)",
            "DPO (days)",
            "FCF (CFO - Capex)",
            "Capex / Revenue %",
        ]:
            if pd.isna(sources.at[ratio_label, year]) or sources.at[ratio_label, year] in (None, ""):
                sources.at[ratio_label, year] = "derived"

    for idx, year in enumerate(years):
        if idx == 0:
            financials.at["Revenue YoY %", year] = None
            financials.at["EBITDA YoY %", year] = None
            sources.at["Revenue YoY %", year] = "n/a"
            sources.at["EBITDA YoY %", year] = "n/a"
            continue
        prev_year = years[idx - 1]
        financials.at["Revenue YoY %", year] = _ratio(financials.at["Revenue", year], financials.at["Revenue", prev_year], fallback=0.0) - 1 if financials.at["Revenue", prev_year] not in (None, 0) else None
        financials.at["EBITDA YoY %", year] = _ratio(financials.at["EBITDA", year], financials.at["EBITDA", prev_year], fallback=0.0) - 1 if financials.at["EBITDA", prev_year] not in (None, 0) else None
        sources.at["Revenue YoY %", year] = "derived"
        sources.at["EBITDA YoY %", year] = "derived"

    if years:
        first_year = years[0]
        if pd.isna(financials.at["Revenue YoY %", first_year]):
            financials.at["Revenue YoY %", first_year] = None
        if pd.isna(financials.at["EBITDA YoY %", first_year]):
            financials.at["EBITDA YoY %", first_year] = None

        latest_year = years[-1]
        latest_interest = financials.at["Interest Expense", latest_year]
        latest_total_debt = (financials.at["Short-term Debt", latest_year] or 0.0) + (financials.at["Long-term Debt", latest_year] or 0.0)
        if _is_missing(latest_interest) and latest_total_debt > 0:
            fallback_interest = None
            fallback_source = None
            for prior_year in reversed(years[:-1]):
                prior_interest = financials.at["Interest Expense", prior_year]
                prior_total_debt = (financials.at["Short-term Debt", prior_year] or 0.0) + (financials.at["Long-term Debt", prior_year] or 0.0)
                if _is_missing(prior_interest) or prior_total_debt <= 0:
                    continue
                implied_rate = prior_interest / prior_total_debt
                fallback_interest = implied_rate * latest_total_debt
                fallback_source = f"fallback: {prior_year} interest rate x latest debt"
                break
            if fallback_interest is not None:
                financials.at["Interest Expense", latest_year] = fallback_interest
                sources.at["Interest Expense", latest_year] = fallback_source
                ebit_latest = financials.at["EBIT", latest_year]
                if not _is_missing(ebit_latest):
                    financials.at["Interest Coverage", latest_year] = _ratio(ebit_latest, fallback_interest, fallback=999.0)
                    sources.at["Interest Coverage", latest_year] = "derived"
                warnings.append(f"{latest_year}: Interest Expense missing in Yahoo annuals; used {fallback_source}.")

    return financials, sources, warnings


def _build_latest_values(financials: pd.DataFrame) -> dict[str, float]:
    latest_year = str(financials.columns[-1])
    latest = {}
    for line_item in financials.index:
        value = financials.at[line_item, latest_year]
        if _is_missing(value):
            continue
        latest[line_item] = float(value)
    latest["Current Assets"] = latest.get("Cash & Equivalents", 0.0) + latest.get("Accounts Receivable", 0.0) + latest.get("Inventory", 0.0) + latest.get("Other Current Assets", 0.0)
    latest["Current Liabilities"] = latest.get("Short-term Debt", 0.0) + latest.get("Accounts Payable", 0.0) + latest.get("Other Current Liabilities", 0.0)
    latest["Total Debt"] = latest.get("Short-term Debt", 0.0) + latest.get("Long-term Debt", 0.0)
    return latest


def prepare_latest_for_stress(latest: dict[str, float]) -> dict[str, float]:
    prepared = dict(latest)

    if _is_missing(prepared.get("Gross Profit")) and not _is_missing(prepared.get("Revenue")) and not _is_missing(prepared.get("COGS")):
        prepared["Gross Profit"] = float(prepared["Revenue"] - prepared["COGS"])

    if _is_missing(prepared.get("EBITDA")) and not _is_missing(prepared.get("Gross Profit")) and not _is_missing(prepared.get("SG&A")) and not _is_missing(prepared.get("Other OpEx")):
        prepared["EBITDA"] = float(prepared["Gross Profit"] - prepared["SG&A"] - prepared["Other OpEx"])

    if _is_missing(prepared.get("EBIT")) and not _is_missing(prepared.get("EBITDA")) and not _is_missing(prepared.get("D&A")):
        prepared["EBIT"] = float(prepared["EBITDA"] - prepared["D&A"])

    if _is_missing(prepared.get("Pretax Income")) and not _is_missing(prepared.get("EBIT")) and not _is_missing(prepared.get("Interest Expense")):
        prepared["Pretax Income"] = float(prepared["EBIT"] - prepared["Interest Expense"])

    if _is_missing(prepared.get("Net Income")) and not _is_missing(prepared.get("Pretax Income")) and not _is_missing(prepared.get("Taxes")):
        prepared["Net Income"] = float(prepared["Pretax Income"] - prepared["Taxes"])

    if _is_missing(prepared.get("CFO (actual, optional)")) and not _is_missing(prepared.get("Net Income")) and not _is_missing(prepared.get("D&A")):
        prepared["CFO (actual, optional)"] = float(prepared["Net Income"] + prepared["D&A"])

    if _is_missing(prepared.get("FCF (CFO - Capex)")) and not _is_missing(prepared.get("CFO (actual, optional)")) and not _is_missing(prepared.get("Capex")):
        prepared["FCF (CFO - Capex)"] = float(prepared["CFO (actual, optional)"] - prepared["Capex"])

    if _is_missing(prepared.get("Net Leverage")) and not _is_missing(prepared.get("EBITDA")):
        prepared["Net Leverage"] = float(
            _ratio(
                prepared.get("Short-term Debt", 0.0) + prepared.get("Long-term Debt", 0.0) - prepared.get("Cash & Equivalents", 0.0),
                prepared["EBITDA"],
                fallback=0.0,
            )
        )

    if _is_missing(prepared.get("Interest Coverage")) and not _is_missing(prepared.get("EBIT")) and not _is_missing(prepared.get("Interest Expense")):
        prepared["Interest Coverage"] = float(_ratio(prepared["EBIT"], prepared["Interest Expense"], fallback=999.0))

    missing_fields = [field for field in STRESS_REQUIRED_FIELDS if _is_missing(prepared.get(field))]
    if missing_fields:
        raise StressModelDataError(missing_fields)
    return prepared


def _compute_data_quality(financials: pd.DataFrame) -> tuple[float, list[str]]:
    latest = _build_latest_values(financials)
    blockers = [field for field in REQUIRED_BASE_FIELDS if field not in latest or latest[field] is None]
    score = 1 - (len(blockers) / len(REQUIRED_BASE_FIELDS))
    return max(0.0, round(score * 100, 1)), blockers


def build_historical_dataset(symbol: str) -> HistoricalDataset:
    annual_raw = fetch_annual_financials(symbol)
    overview = fetch_company_overview(symbol)
    financials, sources, warnings = map_historical_financials(annual_raw)
    latest_year = str(financials.columns[-1])
    latest_values = _build_latest_values(financials)
    data_quality_score, blockers = _compute_data_quality(financials)

    sector_warning = None
    sector_key = overview.sector.lower()
    if "financial" in sector_key or "bank" in overview.industry.lower() or "insurance" in overview.industry.lower():
        sector_warning = (
            "This stress model is designed for operating companies. "
            "Banks, insurers, brokers, and other financial institutions often do not map cleanly to COGS, inventory, and working-capital logic."
        )

    return HistoricalDataset(
        overview=overview,
        annual_raw=annual_raw,
        financials=financials,
        sources=sources,
        latest_year=latest_year,
        latest_values=latest_values,
        blockers=blockers,
        warnings=sorted(set(warnings)),
        data_quality_score=data_quality_score,
        sector_warning=sector_warning,
    )


def run_scenario(latest: dict[str, float], scenario_row: pd.Series, thresholds: ThresholdSettings) -> dict[str, Any]:
    latest = prepare_latest_for_stress(latest)
    revenue_shock = float(scenario_row["Revenue Shock %"])
    gm_shock_bps = float(scenario_row["Gross Margin Shock (bps)"])
    sga_action = float(scenario_row["SG&A Mgmt Action %"])
    other_opex_shock = float(scenario_row["Other OpEx Shock %"])
    dso_change = float(scenario_row["DSO Change"])
    dio_change = float(scenario_row["DIO Change"])
    dpo_change = float(scenario_row["DPO Change"])
    interest_shock_bps = float(scenario_row["Interest Shock (bps)"])
    st_rollover = float(scenario_row["ST Debt Rollover %"])
    lt_due = float(scenario_row["LT Debt Due %"])
    lt_rollover = float(scenario_row["LT Debt Rollover %"])
    capex_cut = float(scenario_row["Capex Cut %"])
    dividend_cut = float(scenario_row["Dividend Cut %"])
    buyback_cut = float(scenario_row["Buyback Cut %"])
    ppe_impairment_pct = float(scenario_row["PP&E Impairment %"])
    intangible_impairment_pct = float(scenario_row["Intangible Impairment %"])
    one_off_charge_pct = float(scenario_row["One-off Cash Charge % Rev"])
    tax_rate = float(scenario_row["Tax Rate %"])

    base_revenue = latest["Revenue"]
    base_cogs = latest["COGS"]
    base_gross_profit = latest["Gross Profit"]
    base_sga = latest["SG&A"]
    base_other_opex = latest["Other OpEx"]
    base_da = latest["D&A"]
    base_interest = latest["Interest Expense"]
    base_cash = latest["Cash & Equivalents"]
    base_receivables = latest["Accounts Receivable"]
    base_inventory = latest["Inventory"]
    base_other_current_assets = latest.get("Other Current Assets", 0.0)
    base_ppe = latest["PP&E, net"]
    base_intangibles = latest.get("Intangibles & Goodwill", 0.0)
    base_short_debt = latest.get("Short-term Debt", 0.0)
    base_payables = latest["Accounts Payable"]
    base_other_current_liabilities = latest.get("Other Current Liabilities", 0.0)
    base_long_debt = latest.get("Long-term Debt", 0.0)
    base_equity = latest["Equity"]
    base_capex = latest["Capex"]
    base_dividends = latest.get("Dividends", 0.0)
    base_buybacks = latest.get("Buybacks", 0.0)
    base_cfo = latest.get("CFO (actual, optional)", latest.get("Net Income", 0.0) + base_da)

    base_total_debt = base_short_debt + base_long_debt
    base_gross_margin = _ratio(base_gross_profit, base_revenue)
    base_dso = _ratio(base_receivables * 365, base_revenue)
    base_dio = _ratio(base_inventory * 365, base_cogs)
    base_dpo = _ratio(base_payables * 365, base_cogs)
    base_interest_rate = _ratio(base_interest, base_total_debt)

    stressed_revenue = base_revenue * (1 + revenue_shock)
    stressed_gross_margin = _clamp(base_gross_margin + gm_shock_bps / 10000, -0.5, 0.9)
    stressed_gross_profit = stressed_revenue * stressed_gross_margin
    stressed_sga = max(0.0, base_sga * (1 + thresholds.sga_variable_cost_share * revenue_shock + sga_action))
    stressed_other_opex = max(0.0, base_other_opex * (1 + other_opex_shock))
    stressed_ebitda = stressed_gross_profit - stressed_sga - stressed_other_opex
    stressed_da = base_da
    stressed_ebit = stressed_ebitda - stressed_da
    stressed_interest_rate = max(0.0, base_interest_rate + interest_shock_bps / 10000)

    mandatory_st_repayment = base_short_debt * (1 - st_rollover)
    mandatory_lt_repayment = base_long_debt * lt_due * (1 - lt_rollover)
    ending_short_debt = base_short_debt - mandatory_st_repayment
    ending_long_debt = base_long_debt - mandatory_lt_repayment
    total_ending_debt = ending_short_debt + ending_long_debt
    average_debt = (base_total_debt + total_ending_debt) / 2
    stressed_interest_expense = stressed_interest_rate * average_debt

    ppe_impairment = base_ppe * ppe_impairment_pct
    intangible_impairment = base_intangibles * intangible_impairment_pct
    one_off_charge = base_revenue * one_off_charge_pct
    stressed_pretax = stressed_ebit - stressed_interest_expense - ppe_impairment - intangible_impairment - one_off_charge
    stressed_taxes = max(0.0, stressed_pretax) * tax_rate
    stressed_net_income = stressed_pretax - stressed_taxes

    stressed_dso = max(0.0, base_dso + dso_change)
    stressed_dio = max(0.0, base_dio + dio_change)
    stressed_dpo = max(0.0, base_dpo + dpo_change)
    stressed_receivables = stressed_revenue * stressed_dso / 365
    stressed_inventory = (stressed_revenue - stressed_gross_profit) * stressed_dio / 365
    stressed_payables = (stressed_revenue - stressed_gross_profit) * stressed_dpo / 365

    stressed_cfo = (
        stressed_net_income
        + stressed_da
        + ppe_impairment
        + intangible_impairment
        - (stressed_receivables - base_receivables)
        - (stressed_inventory - base_inventory)
        + (stressed_payables - base_payables)
    )
    stressed_capex = base_capex * (1 - capex_cut)
    scheduled_dividends = base_dividends * (1 - dividend_cut)
    scheduled_buybacks = base_buybacks * (1 - buyback_cut)
    stressed_dividends = scheduled_dividends
    stressed_buybacks = scheduled_buybacks

    cash_before_financing = base_cash + stressed_cfo - stressed_capex - stressed_dividends - stressed_buybacks
    ending_cash = cash_before_financing - mandatory_st_repayment - mandatory_lt_repayment

    auto_buyback_cut = 0.0
    auto_dividend_cut = 0.0
    if ending_cash < CASH_PRESERVATION_MIN:
        required_cash = CASH_PRESERVATION_MIN - ending_cash
        auto_buyback_cut = min(stressed_buybacks, required_cash)
        stressed_buybacks -= auto_buyback_cut
        required_cash -= auto_buyback_cut
        if required_cash > 0:
            auto_dividend_cut = min(stressed_dividends, required_cash)
            stressed_dividends -= auto_dividend_cut

        cash_before_financing = base_cash + stressed_cfo - stressed_capex - stressed_dividends - stressed_buybacks
        ending_cash = cash_before_financing - mandatory_st_repayment - mandatory_lt_repayment

    funding_gap = max(0.0, thresholds.minimum_cash_buffer - ending_cash)

    ending_equity = base_equity + stressed_net_income - stressed_dividends - stressed_buybacks
    ending_ppe = max(0.0, base_ppe + stressed_capex - stressed_da - ppe_impairment)
    ending_intangibles = max(0.0, base_intangibles - intangible_impairment)

    net_debt_to_ebitda = _ratio(total_ending_debt - ending_cash, stressed_ebitda, fallback=99.0)
    interest_coverage = _ratio(stressed_ebit, stressed_interest_expense, fallback=999.0)
    current_ratio = _ratio(
        max(0.0, ending_cash) + stressed_receivables + stressed_inventory + base_other_current_assets,
        stressed_payables + base_other_current_liabilities + max(0.0, ending_short_debt),
        fallback=0.0,
    )
    fcf = stressed_cfo - stressed_capex

    critical_flags = {
        "Ending cash at or below zero": ending_cash <= DISTRESS_MIN_ENDING_CASH,
        "Net leverage at or above distress ceiling": net_debt_to_ebitda >= DISTRESS_MAX_NET_LEVERAGE,
        "Interest coverage at or below distress floor": interest_coverage <= DISTRESS_MIN_INTEREST_COVERAGE,
        "Current ratio at or below distress floor": current_ratio <= DISTRESS_MIN_CURRENT_RATIO,
        "Ending equity below zero": ending_equity < 0,
        "FCF deeply negative": fcf < (-0.2 * base_revenue),
    }
    benchmark_flags = {
        "Cash below company buffer": ending_cash < thresholds.minimum_cash_buffer,
        "Net leverage above company benchmark": net_debt_to_ebitda > thresholds.maximum_net_leverage,
        "Interest coverage below company benchmark": interest_coverage < thresholds.minimum_interest_coverage,
        "Current ratio below company benchmark": current_ratio < thresholds.minimum_current_ratio,
    }
    watch_flags = {
        "Net leverage at or above watch ceiling": net_debt_to_ebitda >= WATCH_MAX_NET_LEVERAGE,
        "Interest coverage at or below watch floor": interest_coverage <= WATCH_MIN_INTEREST_COVERAGE,
        "Current ratio at or below watch floor": current_ratio <= WATCH_MIN_CURRENT_RATIO,
        "FCF negative": fcf < 0,
    }

    critical_reasons = []
    if critical_flags["Ending cash at or below zero"]:
        critical_reasons.append(f"Ending cash {ending_cash:.1f} <= 0.0")
    if critical_flags["Net leverage at or above distress ceiling"]:
        critical_reasons.append(
            f"Net leverage {net_debt_to_ebitda:.2f}x >= distress ceiling {DISTRESS_MAX_NET_LEVERAGE:.2f}x"
        )
    if critical_flags["Interest coverage at or below distress floor"]:
        critical_reasons.append(
            f"Interest coverage {interest_coverage:.2f}x <= distress floor {DISTRESS_MIN_INTEREST_COVERAGE:.2f}x"
        )
    if critical_flags["Current ratio at or below distress floor"]:
        critical_reasons.append(
            f"Current ratio {current_ratio:.2f}x <= distress floor {DISTRESS_MIN_CURRENT_RATIO:.2f}x"
        )
    if critical_flags["Ending equity below zero"]:
        critical_reasons.append("Ending equity < 0")
    if critical_flags["FCF deeply negative"]:
        critical_reasons.append("FCF < -20% of base revenue")

    benchmark_reasons = []
    if benchmark_flags["Cash below company buffer"]:
        benchmark_reasons.append(
            f"Ending cash {ending_cash:.1f} is below the company buffer {thresholds.minimum_cash_buffer:.1f}"
        )
    if benchmark_flags["Net leverage above company benchmark"]:
        benchmark_reasons.append(
            f"Net leverage {net_debt_to_ebitda:.2f}x > company benchmark {thresholds.maximum_net_leverage:.2f}x"
        )
    if benchmark_flags["Interest coverage below company benchmark"]:
        benchmark_reasons.append(
            f"Interest coverage {interest_coverage:.2f}x < company benchmark {thresholds.minimum_interest_coverage:.2f}x"
        )
    if benchmark_flags["Current ratio below company benchmark"]:
        benchmark_reasons.append(
            f"Current ratio {current_ratio:.2f}x < company benchmark {thresholds.minimum_current_ratio:.2f}x"
        )

    watch_reasons = []
    if auto_buyback_cut > 0:
        watch_reasons.append(f"Buybacks need an automatic cut of {auto_buyback_cut:.1f} to preserve cash")
    if auto_dividend_cut > 0:
        watch_reasons.append(f"Dividends need an automatic cut of {auto_dividend_cut:.1f} to preserve cash")
    if watch_flags["Net leverage at or above watch ceiling"] and not critical_flags["Net leverage at or above distress ceiling"]:
        watch_reasons.append(
            f"Net leverage {net_debt_to_ebitda:.2f}x >= watch ceiling {WATCH_MAX_NET_LEVERAGE:.2f}x"
        )
    if watch_flags["Interest coverage at or below watch floor"] and not critical_flags["Interest coverage at or below distress floor"]:
        watch_reasons.append(
            f"Interest coverage {interest_coverage:.2f}x <= watch floor {WATCH_MIN_INTEREST_COVERAGE:.2f}x"
        )
    if watch_flags["Current ratio at or below watch floor"] and not critical_flags["Current ratio at or below distress floor"]:
        watch_reasons.append(
            f"Current ratio {current_ratio:.2f}x <= watch floor {WATCH_MIN_CURRENT_RATIO:.2f}x"
        )
    if watch_flags["FCF negative"] and not critical_flags["FCF deeply negative"]:
        watch_reasons.append(f"FCF turns negative ({fcf:.1f})")

    if critical_reasons:
        rating = "CRITICAL"
        rating_reasons = critical_reasons
    else:
        has_watch_pressure = any(watch_flags.values())
        rating = "WATCH" if has_watch_pressure else "RESILIENT"
        rating_reasons = watch_reasons if has_watch_pressure else []

    def _min_metric_status(value: float, benchmark_floor: float, distress_floor: float) -> str:
        if value <= distress_floor:
            return "CRITICAL"
        if value < benchmark_floor:
            return "WATCH"
        return "OK"

    def _max_metric_status(value: float, benchmark_ceiling: float, distress_ceiling: float) -> str:
        if value >= distress_ceiling:
            return "CRITICAL"
        if value > benchmark_ceiling:
            return "WATCH"
        return "OK"

    dashboard_status = {
        "Ending cash": "CRITICAL" if ending_cash <= DISTRESS_MIN_ENDING_CASH else "OK",
        "Funding gap": "INFO" if funding_gap > 0 else "OK",
        "Net debt / EBITDA": _max_metric_status(
            net_debt_to_ebitda,
            WATCH_MAX_NET_LEVERAGE,
            DISTRESS_MAX_NET_LEVERAGE,
        ),
        "EBIT / interest": _min_metric_status(
            interest_coverage,
            WATCH_MIN_INTEREST_COVERAGE,
            DISTRESS_MIN_INTEREST_COVERAGE,
        ),
        "Current ratio": _min_metric_status(
            current_ratio,
            WATCH_MIN_CURRENT_RATIO,
            DISTRESS_MIN_CURRENT_RATIO,
        ),
        "FCF": "CRITICAL" if critical_flags["FCF deeply negative"] else "WATCH" if fcf < 0 else "OK",
        "Ending equity": "CRITICAL" if ending_equity < 0 else "OK",
    }

    if rating == "CRITICAL":
        overall_assessment = "Financial distress"
    elif any(watch_flags.values()):
        overall_assessment = "Financial pressure, but above distress floors"
    elif any(benchmark_flags.values()):
        overall_assessment = "Below company benchmark, but outside financial pressure and distress zones"
    else:
        overall_assessment = "Within benchmark and above distress floors"

    return {
        "Sequence": scenario_row["Sequence"],
        "Severity": scenario_row["Severity"],
        "Scenario Key": scenario_row["Scenario Key"],
        "Cause -> Effect Chain": scenario_row["Cause -> Effect Chain"],
        "Revenue (Stressed)": stressed_revenue,
        "Revenue Δ": stressed_revenue - base_revenue,
        "EBITDA Margin (Stressed)": _ratio(stressed_ebitda, stressed_revenue),
        "EBITDA (Stressed)": stressed_ebitda,
        "EBITDA Δ": stressed_ebitda - latest["EBITDA"],
        "Net Income (Stressed)": stressed_net_income,
        "FCF (Stressed)": fcf,
        "Ending Cash": ending_cash,
        "Funding Gap": funding_gap,
        "Net Leverage": net_debt_to_ebitda,
        "Interest Coverage": interest_coverage,
        "Current Ratio": current_ratio,
        "Ending Equity": ending_equity,
        "Rating": rating,
        "Rating Reasons": "; ".join(rating_reasons) if rating_reasons else "No major trigger",
        "Critical Reasons": "; ".join(critical_reasons) if critical_reasons else "None",
        "Benchmark Reasons": "; ".join(benchmark_reasons) if benchmark_reasons else "None",
        "Critical Flags": critical_flags,
        "Benchmark Flags": benchmark_flags,
        "Overall Assessment": overall_assessment,
        "Revenue Shock %": revenue_shock,
        "Gross Margin Shock (bps)": gm_shock_bps,
        "DSO Change": dso_change,
        "DIO Change": dio_change,
        "DPO Change": dpo_change,
        "Interest Shock (bps)": interest_shock_bps,
        "ST Debt Rollover %": st_rollover,
        "LT Debt Due %": lt_due,
        "LT Debt Rollover %": lt_rollover,
        "Capex Cut %": capex_cut,
        "Dividend Cut %": dividend_cut,
        "Buyback Cut %": buyback_cut,
        "One-off Cash Charge % Rev": one_off_charge_pct,
        "Dashboard Status": dashboard_status,
        "Stress Detail": {
            "Base Revenue": base_revenue,
            "Stressed Revenue": stressed_revenue,
            "Base EBITDA": latest["EBITDA"],
            "Stressed EBITDA": stressed_ebitda,
            "Base Net Income": latest["Net Income"],
            "Stressed Net Income": stressed_net_income,
            "Base CFO": base_cfo,
            "Stressed CFO": stressed_cfo,
            "Base Cash": base_cash,
            "Ending Cash": ending_cash,
            "Base Net Leverage": latest["Net Leverage"],
            "Stressed Net Leverage": net_debt_to_ebitda,
            "Base Interest Coverage": latest["Interest Coverage"],
            "Stressed Interest Coverage": interest_coverage,
            "Base Current Ratio": _ratio(
                latest["Cash & Equivalents"] + latest["Accounts Receivable"] + latest["Inventory"] + latest.get("Other Current Assets", 0.0),
                latest.get("Accounts Payable", 0.0) + latest.get("Other Current Liabilities", 0.0) + latest.get("Short-term Debt", 0.0),
                fallback=0.0,
            ),
            "Stressed Current Ratio": current_ratio,
            "Scheduled Dividends": scheduled_dividends,
            "Actual Dividends": stressed_dividends,
            "Scheduled Buybacks": scheduled_buybacks,
            "Actual Buybacks": stressed_buybacks,
            "Automatic Buyback Cut": auto_buyback_cut,
            "Automatic Dividend Cut": auto_dividend_cut,
            "One-off Cash Charge": one_off_charge,
            "PP&E Impairment": ppe_impairment,
            "Intangible Impairment": intangible_impairment,
        },
    }


def run_all_scenarios(latest: dict[str, float], thresholds: ThresholdSettings, scenario_library: pd.DataFrame) -> pd.DataFrame:
    prepared_latest = prepare_latest_for_stress(latest)
    records = [run_scenario(prepared_latest, row, thresholds) for _, row in scenario_library.iterrows()]
    matrix = pd.DataFrame(records)
    severity_rank = {"Light": 1, "Base": 2, "Severe": 3}
    rating_rank = {"CRITICAL": 1, "WATCH": 2, "RESILIENT": 3}
    matrix["Severity Rank"] = matrix["Severity"].map(severity_rank).fillna(99)
    matrix["Rating Rank"] = matrix["Rating"].map(rating_rank).fillna(99)
    matrix = matrix.sort_values(["Severity Rank", "Rating Rank", "Sequence"]).reset_index(drop=True)
    return matrix


def get_selected_scenario(scenario_library: pd.DataFrame, sequence: str, severity: str) -> pd.Series:
    matched = scenario_library[(scenario_library["Sequence"] == sequence) & (scenario_library["Severity"] == severity)]
    if matched.empty:
        raise ValueError(f"Scenario '{sequence} | {severity}' was not found in the workbook library.")
    return matched.iloc[0]
