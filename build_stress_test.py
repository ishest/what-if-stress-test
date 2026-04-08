from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_FILE = "stress_test_v1.xlsx"

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="EEF3F7")
ASSUMPTION_FILL = PatternFill("solid", fgColor="E2F0D9")
STRESS_FILL = PatternFill("solid", fgColor="FCE4D6")
OUTPUT_FILL = PatternFill("solid", fgColor="DDEBF7")
PASS_FILL = PatternFill("solid", fgColor="C6E0B4")
FAIL_FILL = PatternFill("solid", fgColor="F8CBAD")
INFO_FILL = PatternFill("solid", fgColor="E7E6E6")

THIN = Side(style="thin", color="B7C9D6")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SEVERITY_ORDER = ["Light", "Base", "Severe"]


def scenario_values(
    *,
    rev=0.0,
    gm=0,
    sga=0.0,
    opex=0.0,
    dso=0,
    dio=0,
    dpo=0,
    rate=0,
    st_roll=1.0,
    lt_due=0.08,
    lt_roll=1.0,
    capex=0.0,
    div=0.0,
    buy=0.0,
    ppe=0.0,
    intg=0.0,
    charge=0.0,
    tax=0.25,
):
    return {
        "rev": rev,
        "gm": gm,
        "sga": sga,
        "opex": opex,
        "dso": dso,
        "dio": dio,
        "dpo": dpo,
        "rate": rate,
        "st_roll": st_roll,
        "lt_due": lt_due,
        "lt_roll": lt_roll,
        "capex": capex,
        "div": div,
        "buy": buy,
        "ppe": ppe,
        "intg": intg,
        "charge": charge,
        "tax": tax,
    }


SEQUENCE_DEFINITIONS = [
    {
        "name": "Custom / Blank",
        "chain": "Zero-shock starting point -> override the drivers manually in Scenario_Setup.",
        "driver_inputs": "Any driver in Scenario_Setup.",
        "outputs": "Use all model outputs as a free-form what-if.",
        "note": "Useful for screening your own thesis or management guidance downside.",
        "severities": {
            "Light": scenario_values(),
            "Base": scenario_values(),
            "Severe": scenario_values(),
        },
    },
    {
        "name": "Demand collapse (classic recession)",
        "chain": "Revenue falls -> fixed costs stay sticky -> EBITDA drops hard -> leverage spikes -> refinancing gets harder -> liquidity stress builds.",
        "driver_inputs": "Revenue, gross margin, SG&A flexibility, DSO/DIO/DPO, rates, debt rollover, payout cuts.",
        "outputs": "EBITDA, CFO, ending cash, funding gap, net debt / EBITDA, EBIT / interest.",
        "note": "Closest standard template for a GM 2008-style demand shock.",
        "severities": {
            "Light": scenario_values(rev=-0.10, gm=-150, sga=-0.01, dso=5, dio=8, dpo=-2, rate=100, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.10, div=0.00, buy=0.50, tax=0.22),
            "Base": scenario_values(rev=-0.20, gm=-300, sga=-0.02, dso=10, dio=15, dpo=-3, rate=200, st_roll=0.85, lt_due=0.10, lt_roll=0.80, capex=0.25, div=0.50, buy=1.00, ppe=0.02, intg=0.05, tax=0.20),
            "Severe": scenario_values(rev=-0.30, gm=-500, sga=-0.03, dso=15, dio=20, dpo=-5, rate=300, st_roll=0.60, lt_due=0.15, lt_roll=0.60, capex=0.40, div=1.00, buy=1.00, ppe=0.05, intg=0.10, tax=0.18),
        },
    },
    {
        "name": "Mild demand shock + operating leverage",
        "chain": "Revenue dips -> fixed cost absorption weakens -> margins compress -> leverage rises -> rating and funding cost worsen.",
        "driver_inputs": "Revenue, gross margin, SG&A flexibility, rates, payout defense.",
        "outputs": "EBITDA margin, net debt / EBITDA, interest coverage, equity value pressure proxies.",
        "note": "Useful for asking what combination of smaller changes breaks the company.",
        "severities": {
            "Light": scenario_values(rev=-0.05, gm=-100, sga=0.00, dso=2, dio=3, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.05, tax=0.23),
            "Base": scenario_values(rev=-0.10, gm=-300, sga=0.00, dso=4, dio=5, dpo=-1, rate=100, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.10, div=0.25, buy=0.75, tax=0.22),
            "Severe": scenario_values(rev=-0.15, gm=-450, sga=-0.01, dso=6, dio=8, dpo=-2, rate=150, st_roll=0.95, lt_due=0.10, lt_roll=0.90, capex=0.15, div=0.50, buy=1.00, tax=0.20),
        },
    },
    {
        "name": "Price pressure + volume decline",
        "chain": "Volume falls and prices weaken -> revenue drops -> gross margin shrinks -> FCF turns negative -> debt burden rises.",
        "driver_inputs": "Revenue, gross margin, DSO/DIO/DPO, rates, rollover terms.",
        "outputs": "Revenue, EBITDA, CFO, debt build, refinancing risk.",
        "note": "Good for commoditized sectors with weak pricing power.",
        "severities": {
            "Light": scenario_values(rev=-0.10, gm=-150, sga=-0.01, dso=3, dio=5, dpo=-1, rate=75, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.05, tax=0.22),
            "Base": scenario_values(rev=-0.20, gm=-250, sga=-0.02, dso=5, dio=10, dpo=-2, rate=125, st_roll=0.95, lt_due=0.10, lt_roll=0.90, capex=0.15, div=0.25, buy=0.75, tax=0.21),
            "Severe": scenario_values(rev=-0.25, gm=-400, sga=-0.02, dso=8, dio=15, dpo=-4, rate=200, st_roll=0.90, lt_due=0.12, lt_roll=0.80, capex=0.25, div=0.50, buy=1.00, ppe=0.02, tax=0.19),
        },
    },
    {
        "name": "Customer concentration risk",
        "chain": "Top customer is lost -> revenue gap opens immediately -> cost base stays in place -> liquidity tightens -> forced actions follow.",
        "driver_inputs": "Revenue, SG&A stickiness, DSO/DIO unwind, rates, debt rollover.",
        "outputs": "Revenue gap, EBITDA collapse, cash drain, funding gap, asset-sale pressure.",
        "note": "Best paired with a realistic minimum cash buffer in Scenario_Setup.",
        "severities": {
            "Light": scenario_values(rev=-0.10, gm=-100, sga=0.00, dso=4, dio=5, dpo=-1, rate=75, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.10, div=0.00, buy=1.00, tax=0.22),
            "Base": scenario_values(rev=-0.20, gm=-200, sga=0.00, dso=8, dio=10, dpo=-3, rate=150, st_roll=0.90, lt_due=0.10, lt_roll=0.80, capex=0.20, div=0.50, buy=1.00, intg=0.03, tax=0.20),
            "Severe": scenario_values(rev=-0.30, gm=-300, sga=-0.01, dso=12, dio=15, dpo=-5, rate=250, st_roll=0.75, lt_due=0.12, lt_roll=0.65, capex=0.30, div=1.00, buy=1.00, ppe=0.03, intg=0.08, tax=0.18),
        },
    },
    {
        "name": "Input cost inflation shock",
        "chain": "COGS rises -> pricing lags -> gross margin falls -> interest coverage tightens -> covenant headroom disappears.",
        "driver_inputs": "Gross margin, other OpEx, inventory days, rates, payout defense.",
        "outputs": "Gross profit, EBITDA, coverage ratio, covenant headroom proxies.",
        "note": "Useful for importers, processors, and industries with delayed pricing pass-through.",
        "severities": {
            "Light": scenario_values(rev=0.00, gm=-250, opex=0.01, dio=5, dpo=1, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.05, tax=0.24),
            "Base": scenario_values(rev=-0.02, gm=-400, opex=0.02, dio=10, dpo=2, rate=100, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.10, div=0.25, buy=0.75, tax=0.22),
            "Severe": scenario_values(rev=-0.05, gm=-600, opex=0.03, dio=15, dpo=3, rate=150, st_roll=0.95, lt_due=0.10, lt_roll=0.90, capex=0.15, div=0.50, buy=1.00, tax=0.20),
        },
    },
    {
        "name": "Wage inflation + sticky pricing",
        "chain": "Labor costs rise while pricing stays flat -> EBITDA margin compresses -> FCF falls -> management slows hiring and growth.",
        "driver_inputs": "SG&A inflation, other OpEx, capex response, payout cuts.",
        "outputs": "EBITDA margin, FCF, growth capacity, medium-term earnings power.",
        "note": "Better for service-heavy companies where labor sits outside COGS.",
        "severities": {
            "Light": scenario_values(rev=0.00, gm=0, sga=0.03, opex=0.01, rate=25, st_roll=1.00, lt_due=0.08, lt_roll=1.00, capex=0.05, tax=0.24),
            "Base": scenario_values(rev=0.00, gm=0, sga=0.06, opex=0.02, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.10, div=0.25, buy=0.50, tax=0.23),
            "Severe": scenario_values(rev=-0.03, gm=-50, sga=0.09, opex=0.03, rate=75, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.15, div=0.50, buy=1.00, tax=0.21),
        },
    },
    {
        "name": "Working capital trap",
        "chain": "Receivables and inventory rise faster than payables -> cash gets locked in operations -> revolver need rises -> lenders gain leverage.",
        "driver_inputs": "DSO, DIO, DPO, rates, rollover assumptions.",
        "outputs": "Modeled CFO, ending cash, funding gap, leverage.",
        "note": "One of the best practical screens for a fragile balance sheet.",
        "severities": {
            "Light": scenario_values(rev=-0.03, gm=-50, dso=10, dio=8, dpo=0, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.05, tax=0.23),
            "Base": scenario_values(rev=-0.05, gm=-100, dso=20, dio=15, dpo=-2, rate=100, st_roll=0.95, lt_due=0.08, lt_roll=0.95, capex=0.10, div=0.25, buy=0.75, tax=0.22),
            "Severe": scenario_values(rev=-0.08, gm=-150, dso=30, dio=25, dpo=-5, rate=150, st_roll=0.85, lt_due=0.10, lt_roll=0.85, capex=0.20, div=0.50, buy=1.00, tax=0.20),
        },
    },
    {
        "name": "Inventory overbuild + demand slowdown",
        "chain": "Inventory builds while demand softens -> discounting starts -> margins fall -> write-down pressure and cash burn accelerate.",
        "driver_inputs": "Revenue, gross margin, DIO, DSO/DPO, other OpEx.",
        "outputs": "Gross profit, CFO, ending cash, liquidity squeeze.",
        "note": "Very useful for retailers, auto, industrial distributors, and cyclical manufacturers.",
        "severities": {
            "Light": scenario_values(rev=-0.05, gm=-150, opex=0.01, dso=2, dio=20, dpo=-2, rate=75, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.05, tax=0.22),
            "Base": scenario_values(rev=-0.10, gm=-250, opex=0.02, dso=5, dio=35, dpo=-3, rate=125, st_roll=0.95, lt_due=0.10, lt_roll=0.90, capex=0.10, div=0.25, buy=0.75, tax=0.20),
            "Severe": scenario_values(rev=-0.15, gm=-350, opex=0.03, dso=8, dio=50, dpo=-5, rate=200, st_roll=0.85, lt_due=0.12, lt_roll=0.75, capex=0.20, div=0.50, buy=1.00, ppe=0.02, tax=0.18),
        },
    },
    {
        "name": "Capex overrun + weak returns",
        "chain": "Capex rises before returns show up -> FCF goes negative -> debt grows -> ROIC and valuation compress.",
        "driver_inputs": "Capex increase, modest revenue drag, rates, impairment risk.",
        "outputs": "FCF, ending debt, ending cash, equity dilution risk proxy.",
        "note": "Negative capex cut values mean capex goes up rather than down.",
        "severities": {
            "Light": scenario_values(rev=0.00, gm=0, dso=1, dio=3, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=-0.10, div=0.25, buy=0.50, ppe=0.02, tax=0.24),
            "Base": scenario_values(rev=-0.02, gm=-50, dso=2, dio=5, rate=100, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=-0.20, div=0.50, buy=1.00, ppe=0.05, tax=0.22),
            "Severe": scenario_values(rev=-0.05, gm=-100, dso=5, dio=10, dpo=-1, rate=150, st_roll=0.95, lt_due=0.10, lt_roll=0.85, capex=-0.30, div=1.00, buy=1.00, ppe=0.08, tax=0.20),
        },
    },
    {
        "name": "Refinancing shock",
        "chain": "Debt maturity wall meets tighter credit -> interest cost jumps -> coverage falls -> distressed refinancing or dilution follows.",
        "driver_inputs": "Interest rate shock, ST/LT rollover, LT debt due, payout defense.",
        "outputs": "Interest expense, coverage, ending cash, funding gap, distress risk.",
        "note": "Set minimum cash buffer to the actual revolver cushion or covenant cash minimum.",
        "severities": {
            "Light": scenario_values(rev=0.00, gm=0, rate=150, st_roll=0.80, lt_due=0.10, lt_roll=0.75, capex=0.10, div=0.50, buy=1.00, tax=0.23),
            "Base": scenario_values(rev=-0.03, gm=-50, dso=2, dio=2, rate=300, st_roll=0.50, lt_due=0.15, lt_roll=0.50, capex=0.25, div=1.00, buy=1.00, tax=0.21),
            "Severe": scenario_values(rev=-0.08, gm=-100, dso=5, dio=5, dpo=-1, rate=500, st_roll=0.20, lt_due=0.20, lt_roll=0.20, capex=0.40, div=1.00, buy=1.00, charge=0.01, tax=0.18),
        },
    },
    {
        "name": "Rising rates + high leverage",
        "chain": "Floating-rate debt reprices upward -> interest burden grows -> FCF falls -> leverage rises -> rating pressure increases.",
        "driver_inputs": "Interest rate shock, rollover assumptions, modest capex or payout response.",
        "outputs": "Interest expense, EBIT / interest, net debt / EBITDA, equity downside.",
        "note": "Best for companies that already start with elevated leverage.",
        "severities": {
            "Light": scenario_values(rev=0.00, gm=0, rate=150, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.00, tax=0.24),
            "Base": scenario_values(rev=0.00, gm=0, rate=250, st_roll=0.98, lt_due=0.10, lt_roll=0.90, capex=0.10, div=0.25, buy=0.50, tax=0.23),
            "Severe": scenario_values(rev=-0.03, gm=-50, rate=400, st_roll=0.95, lt_due=0.12, lt_roll=0.75, capex=0.20, div=0.50, buy=1.00, tax=0.21),
        },
    },
    {
        "name": "FX shock (import-heavy company)",
        "chain": "Local currency weakens -> import costs rise -> margin shrinks -> pricing lags -> profitability deteriorates.",
        "driver_inputs": "Gross margin, inventory days, modest revenue drag, rates.",
        "outputs": "Gross profit, EBITDA, cash conversion, coverage.",
        "note": "Useful where raw materials or finished goods are sourced in hard currency.",
        "severities": {
            "Light": scenario_values(rev=0.00, gm=-150, dio=5, dpo=1, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.00, tax=0.24),
            "Base": scenario_values(rev=-0.03, gm=-250, dio=10, dpo=2, rate=100, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.05, tax=0.22),
            "Severe": scenario_values(rev=-0.05, gm=-400, dio=15, dpo=3, rate=150, st_roll=0.95, lt_due=0.10, lt_roll=0.90, capex=0.10, div=0.25, buy=0.50, tax=0.20),
        },
    },
    {
        "name": "FX shock (export-heavy company)",
        "chain": "Currency strengthens -> export competitiveness falls -> utilization weakens -> revenue and margins drift lower.",
        "driver_inputs": "Revenue, gross margin, modest working-capital drag, payout defense.",
        "outputs": "Revenue, EBITDA, cash generation, valuation compression proxies.",
        "note": "Helpful for exporters with high fixed manufacturing footprints.",
        "severities": {
            "Light": scenario_values(rev=-0.05, gm=-100, dso=2, dio=2, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.05, tax=0.23),
            "Base": scenario_values(rev=-0.10, gm=-200, dso=4, dio=5, dpo=-1, rate=75, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.10, div=0.25, buy=0.50, tax=0.22),
            "Severe": scenario_values(rev=-0.15, gm=-300, dso=6, dio=10, dpo=-2, rate=100, st_roll=0.95, lt_due=0.10, lt_roll=0.90, capex=0.15, div=0.50, buy=1.00, tax=0.20),
        },
    },
    {
        "name": "Regulatory shock",
        "chain": "New regulation arrives -> compliance cost and redesign spending rise -> capex and opex increase -> margins and growth slow.",
        "driver_inputs": "Other OpEx, capex increase, modest revenue/gross-margin drag, rates.",
        "outputs": "EBITDA, FCF, project delay pressure, long-term cash drag.",
        "note": "Negative capex cut values represent higher required compliance capex.",
        "severities": {
            "Light": scenario_values(rev=-0.02, gm=-50, opex=0.03, dso=1, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=-0.05, tax=0.23),
            "Base": scenario_values(rev=-0.05, gm=-150, opex=0.06, dso=2, rate=100, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=-0.10, div=0.25, buy=0.50, tax=0.21),
            "Severe": scenario_values(rev=-0.10, gm=-250, opex=0.10, dso=5, dio=3, dpo=-1, rate=150, st_roll=0.95, lt_due=0.10, lt_roll=0.85, capex=-0.15, div=0.50, buy=1.00, intg=0.03, tax=0.18),
        },
    },
    {
        "name": "Litigation / fine event",
        "chain": "Large fine or settlement hits -> cash falls -> leverage rises -> rating and refinancing pressure increase.",
        "driver_inputs": "One-off cash charge, rates, rollover assumptions, payout defense.",
        "outputs": "Net income, CFO, ending cash, funding gap, refinancing flexibility.",
        "note": "The one-off cash charge is modeled as a percent of revenue for portability across companies.",
        "severities": {
            "Light": scenario_values(rev=0.00, gm=0, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.00, charge=0.02, tax=0.24),
            "Base": scenario_values(rev=-0.02, gm=0, rate=100, st_roll=0.95, lt_due=0.10, lt_roll=0.90, capex=0.10, div=0.25, buy=0.50, charge=0.05, tax=0.22),
            "Severe": scenario_values(rev=-0.05, gm=-50, rate=150, st_roll=0.90, lt_due=0.12, lt_roll=0.75, capex=0.20, div=0.50, buy=1.00, charge=0.10, intg=0.03, tax=0.20),
        },
    },
    {
        "name": "Operational disruption",
        "chain": "Plant shutdown or supply disruption hits production -> revenue falls while fixed costs stay -> margin collapses -> cash burn follows.",
        "driver_inputs": "Revenue, gross margin, other OpEx, DSO/DIO, one-off charge, rates.",
        "outputs": "EBITDA, CFO, ending cash, short-term liquidity stress.",
        "note": "Good for accidents, outages, or severe supply chain interruptions.",
        "severities": {
            "Light": scenario_values(rev=-0.08, gm=-150, opex=0.01, dso=2, dio=5, rate=75, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.00, charge=0.01, tax=0.22),
            "Base": scenario_values(rev=-0.15, gm=-300, opex=0.03, dso=5, dio=10, dpo=-2, rate=150, st_roll=0.90, lt_due=0.10, lt_roll=0.85, capex=0.10, div=0.25, buy=0.75, charge=0.03, tax=0.20),
            "Severe": scenario_values(rev=-0.25, gm=-500, opex=0.05, dso=8, dio=15, dpo=-4, rate=200, st_roll=0.80, lt_due=0.12, lt_roll=0.70, capex=0.20, div=0.50, buy=1.00, charge=0.05, ppe=0.03, tax=0.18),
        },
    },
    {
        "name": "Supplier failure",
        "chain": "Key supplier fails -> input shortage cuts production -> urgent sourcing costs more -> margins and cash generation weaken.",
        "driver_inputs": "Revenue, gross margin, other OpEx, DIO, DSO, rates.",
        "outputs": "Revenue, EBITDA, working-capital strain, cash burn.",
        "note": "Useful where single-source components matter.",
        "severities": {
            "Light": scenario_values(rev=-0.05, gm=-100, opex=0.01, dso=1, dio=5, rate=50, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.00, tax=0.22),
            "Base": scenario_values(rev=-0.10, gm=-200, opex=0.02, dso=3, dio=10, dpo=-2, rate=100, st_roll=0.95, lt_due=0.10, lt_roll=0.90, capex=0.05, div=0.25, buy=0.50, tax=0.20),
            "Severe": scenario_values(rev=-0.20, gm=-350, opex=0.04, dso=5, dio=20, dpo=-4, rate=150, st_roll=0.85, lt_due=0.12, lt_roll=0.75, capex=0.10, div=0.50, buy=1.00, charge=0.02, tax=0.18),
        },
    },
    {
        "name": "Technology disruption",
        "chain": "New technology changes the market -> pricing pressure rises -> share is lost -> earnings power erodes over time.",
        "driver_inputs": "Revenue, gross margin, SG&A / OpEx investment, impairment risk.",
        "outputs": "Revenue, EBITDA margin, long-run earnings decline, intangible write-down risk.",
        "note": "Works well for mature incumbents facing a new platform or product substitute.",
        "severities": {
            "Light": scenario_values(rev=-0.05, gm=-100, sga=0.02, opex=0.01, rate=50, st_roll=1.00, lt_due=0.08, lt_roll=0.98, capex=0.00, intg=0.02, tax=0.23),
            "Base": scenario_values(rev=-0.10, gm=-200, sga=0.04, opex=0.02, rate=100, st_roll=0.98, lt_due=0.08, lt_roll=0.95, capex=0.00, div=0.25, buy=0.50, intg=0.05, tax=0.21),
            "Severe": scenario_values(rev=-0.20, gm=-350, sga=0.06, opex=0.03, dso=2, dio=3, rate=150, st_roll=0.95, lt_due=0.10, lt_roll=0.85, capex=0.05, div=0.50, buy=1.00, intg=0.10, tax=0.18),
        },
    },
    {
        "name": "Acquisition gone wrong",
        "chain": "Debt-funded deal underdelivers -> synergies fail -> leverage stays high -> goodwill gets impaired -> refinancing risk rises.",
        "driver_inputs": "Gross margin, other OpEx, DSO/DIO, rates, rollover, goodwill impairment.",
        "outputs": "EBITDA, net debt / EBITDA, ending equity, refinancing risk.",
        "note": "For a real post-deal case, also update starting debt and goodwill in Hist_Financials.",
        "severities": {
            "Light": scenario_values(rev=0.00, gm=-50, opex=0.02, dso=2, dio=2, rate=100, st_roll=0.98, lt_due=0.08, lt_roll=0.90, capex=0.05, div=0.25, buy=1.00, intg=0.08, tax=0.22),
            "Base": scenario_values(rev=-0.05, gm=-100, opex=0.04, dso=5, dio=5, dpo=-1, rate=150, st_roll=0.90, lt_due=0.10, lt_roll=0.75, capex=0.10, div=0.50, buy=1.00, ppe=0.02, intg=0.15, charge=0.01, tax=0.20),
            "Severe": scenario_values(rev=-0.10, gm=-200, opex=0.06, dso=8, dio=8, dpo=-3, rate=250, st_roll=0.75, lt_due=0.12, lt_roll=0.60, capex=0.20, div=1.00, buy=1.00, ppe=0.05, intg=0.25, charge=0.02, tax=0.18),
        },
    },
    {
        "name": "Full death spiral",
        "chain": "Revenue falls -> margin compresses -> working capital drains cash -> rates rise -> refinancing gets expensive -> liquidity is exhausted.",
        "driver_inputs": "Revenue, gross margin, DSO/DIO/DPO, rates, rollover, impairments, one-off charge, payout defense.",
        "outputs": "Every distress output: EBITDA, CFO, ending cash, funding gap, leverage, coverage, negative equity risk.",
        "note": "Use this as a high-end failure mode rather than a base case.",
        "severities": {
            "Light": scenario_values(rev=-0.10, gm=-200, sga=-0.01, dso=10, dio=10, dpo=-2, rate=200, st_roll=0.80, lt_due=0.10, lt_roll=0.80, capex=0.20, div=0.50, buy=1.00, ppe=0.02, intg=0.05, tax=0.20),
            "Base": scenario_values(rev=-0.20, gm=-400, sga=-0.02, dso=20, dio=20, dpo=-5, rate=350, st_roll=0.50, lt_due=0.15, lt_roll=0.60, capex=0.40, div=1.00, buy=1.00, ppe=0.08, intg=0.15, charge=0.03, tax=0.18),
            "Severe": scenario_values(rev=-0.30, gm=-600, sga=-0.03, dso=30, dio=30, dpo=-8, rate=500, st_roll=0.20, lt_due=0.20, lt_roll=0.30, capex=0.50, div=1.00, buy=1.00, ppe=0.15, intg=0.25, charge=0.05, tax=0.15),
        },
    },
]


PARAMETER_MAP = [
    ("Revenue shock %", "%", "E"),
    ("Gross margin shock", "bps", "F"),
    ("SG&A management action %", "%", "G"),
    ("Other OpEx shock %", "%", "H"),
    ("DSO change", "days", "I"),
    ("DIO change", "days", "J"),
    ("DPO change", "days", "K"),
    ("Interest rate shock", "bps", "L"),
    ("Short-term debt rollover %", "%", "M"),
    ("Long-term debt due %", "%", "N"),
    ("Long-term debt rollover %", "%", "O"),
    ("Capex cut %", "%", "P"),
    ("Dividend cut %", "%", "Q"),
    ("Buyback cut %", "%", "R"),
    ("PP&E impairment %", "%", "S"),
    ("Intangible impairment %", "%", "T"),
    ("One-off cash charge % of revenue", "%", "U"),
    ("Tax rate used %", "%", "V"),
]


def style_cell(cell, *, fill=None, bold=False, num_format=None, align="left", font_color=None, wrap=False):
    cell.border = BOX
    if fill:
        cell.fill = fill
    if num_format:
        cell.number_format = num_format
    cell.font = Font(bold=bold, color=font_color if font_color else "000000")
    horizontal = "left"
    if align == "center":
        horizontal = "center"
    elif align == "right":
        horizontal = "right"
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def merge_title(ws, title, cell_range):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    style_cell(cell, fill=TITLE_FILL, bold=True, font_color="FFFFFF", align="center")
    ws.row_dimensions[cell.row].height = 24


def autofit_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def fmt_pct(value):
    return f"{value:+.0%}"


def summarize_capex(capex_cut):
    if capex_cut == 0:
        return ""
    if capex_cut > 0:
        return f"Capex -{abs(capex_cut):.0%}"
    return f"Capex +{abs(capex_cut):.0%}"


def summarize_rollover(st_roll, lt_roll):
    if st_roll >= 0.999 and lt_roll >= 0.999:
        return ""
    return f"Roll ST {st_roll:.0%} / LT {lt_roll:.0%}"


def summarize_severity(values):
    if all(
        [
            values["rev"] == 0,
            values["gm"] == 0,
            values["sga"] == 0,
            values["opex"] == 0,
            values["dso"] == 0,
            values["dio"] == 0,
            values["dpo"] == 0,
            values["rate"] == 0,
            values["st_roll"] == 1,
            values["lt_roll"] == 1,
            values["capex"] == 0,
            values["div"] == 0,
            values["buy"] == 0,
            values["ppe"] == 0,
            values["intg"] == 0,
            values["charge"] == 0,
        ]
    ):
        return "Zero shock; override any driver manually."

    parts = []
    if values["rev"] != 0:
        parts.append(f"Rev {fmt_pct(values['rev'])}")
    if values["gm"] != 0:
        parts.append(f"GM {values['gm']:+.0f} bps")
    if values["sga"] != 0:
        parts.append(f"SG&A {fmt_pct(values['sga'])}")
    if values["opex"] != 0:
        parts.append(f"OpEx {fmt_pct(values['opex'])}")
    if values["dso"] != 0 or values["dio"] != 0 or values["dpo"] != 0:
        parts.append(f"WC DSO {values['dso']:+.0f}d / DIO {values['dio']:+.0f}d / DPO {values['dpo']:+.0f}d")
    if values["rate"] != 0:
        parts.append(f"Rates {values['rate']:+.0f} bps")
    capex_text = summarize_capex(values["capex"])
    if capex_text:
        parts.append(capex_text)
    rollover_text = summarize_rollover(values["st_roll"], values["lt_roll"])
    if rollover_text:
        parts.append(rollover_text)
    if values["charge"] != 0:
        parts.append(f"Cash charge {values['charge']:.0%} of revenue")
    return "; ".join(parts[:5])


def build_scenario_rows():
    rows = []
    for sequence in SEQUENCE_DEFINITIONS:
        for severity in SEVERITY_ORDER:
            values = sequence["severities"][severity]
            rows.append(
                {
                    "sequence": sequence["name"],
                    "severity": severity,
                    "key": f"{sequence['name']} | {severity}",
                    "chain": sequence["chain"],
                    "values": values,
                }
            )
    return rows


SCENARIO_ROWS = build_scenario_rows()
WHATIF_LAST_ROW = 3 + len(SEQUENCE_DEFINITIONS)
SCENARIO_LAST_ROW = 3 + len(SCENARIO_ROWS)


def add_readme(ws):
    merge_title(ws, '"What If..." Company Stress Test', "A1:H1")
    rows = [
        ("Purpose", "Stress test a single company using only income statement, balance sheet, and cash flow data."),
        ("Workflow", "1) Enter three historical years in Hist_Financials. 2) Pick a sequence and severity in Scenario_Setup. 3) Override any driver if needed. 4) Review Stress_Model and Dashboard."),
        ("What-if library", "The workbook now includes 20 cause -> effect sequences plus Light, Base, and Severe versions for each one."),
        ("Key design", "Each sequence maps into driver inputs -> statement impact -> outputs. The focus is on the combinations of smaller changes that break the company."),
        ("GM / 2008 use", "Start with Demand collapse (classic recession) or Full death spiral. Those are the closest generic templates for a 2008-style crisis."),
        ("Core outputs", "Revenue, EBITDA, EBIT, net income, modeled CFO, ending cash, funding gap, net debt / EBITDA, EBIT / interest, current ratio, and ending equity."),
        ("New driver", "One-off cash charge % of revenue lets you model litigation, fines, recalls, or other cash events that hit earnings and liquidity together."),
        ("Capex note", "Capex cut % can be negative. A negative value means capex increases, which is useful for capex overruns or compliance projects."),
        ("Limitations", "This is a fast screening framework. It does not capture segment detail, pension liabilities, covenant definitions, off-balance-sheet exposures, or equity market reflexivity."),
    ]

    ws["A3"] = "How To Use"
    ws["B3"] = "Notes"
    style_cell(ws["A3"], fill=SECTION_FILL, bold=True)
    style_cell(ws["B3"], fill=SECTION_FILL, bold=True)

    row = 4
    for label, text in rows:
        ws[f"A{row}"] = label
        style_cell(ws[f"A{row}"], fill=OUTPUT_FILL, bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws[f"B{row}"] = text
        style_cell(ws[f"B{row}"], fill=FORMULA_FILL, wrap=True)
        ws.row_dimensions[row].height = 34
        row += 1

    ws["A14"] = "What usually breaks first"
    style_cell(ws["A14"], fill=SECTION_FILL, bold=True)
    insight_rows = [
        "Liquidity: ending cash drops below the minimum cash buffer or the funding gap turns positive.",
        "Solvency: net debt / EBITDA breaches the threshold, EBIT / interest falls below the threshold, or equity turns negative.",
        "Cash conversion: DSO and DIO rise while DPO does not offset the pressure.",
        "Management defense: capex, dividends, and buybacks can be cut, but sometimes the operating hit is still too large.",
    ]
    for idx, text in enumerate(insight_rows, start=15):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
        ws[f"A{idx}"] = "- " + text
        style_cell(ws[f"A{idx}"], fill=FORMULA_FILL, wrap=True)

    autofit_widths(ws, {"A": 22, "B": 22, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14})


def add_hist_financials(ws):
    merge_title(ws, "Historical Financials Input", "A1:E1")
    ws["A2"] = "Enter three historical years in a consistent currency and unit. Yellow cells are manual inputs."
    ws.merge_cells("A2:E2")
    style_cell(ws["A2"], fill=FORMULA_FILL)

    headers = ["Line Item", 2023, 2024, 2025, "Comment"]
    for col_idx, value in enumerate(headers, start=1):
        style_cell(ws.cell(row=3, column=col_idx, value=value), fill=SECTION_FILL, bold=True, align="center")

    row_defs = [
        (5, "Income Statement", "section"),
        (6, "Revenue", "input"),
        (7, "COGS", "input"),
        (8, "Gross Profit", "formula", lambda c: f"={c}6-{c}7"),
        (9, "SG&A", "input"),
        (10, "Other OpEx", "input"),
        (11, "EBITDA", "formula", lambda c: f"={c}8-{c}9-{c}10"),
        (12, "D&A", "input"),
        (13, "EBIT", "formula", lambda c: f"={c}11-{c}12"),
        (14, "Interest Expense", "input"),
        (15, "Pretax Income", "formula", lambda c: f"={c}13-{c}14"),
        (16, "Taxes", "input"),
        (17, "Net Income", "formula", lambda c: f"={c}15-{c}16"),
        (19, "Balance Sheet", "section"),
        (20, "Cash & Equivalents", "input"),
        (21, "Accounts Receivable", "input"),
        (22, "Inventory", "input"),
        (23, "Other Current Assets", "input"),
        (24, "PP&E, net", "input"),
        (25, "Intangibles & Goodwill", "input"),
        (26, "Other Non-current Assets", "input"),
        (27, "Short-term Debt", "input"),
        (28, "Accounts Payable", "input"),
        (29, "Other Current Liabilities", "input"),
        (30, "Long-term Debt", "input"),
        (31, "Other Non-current Liabilities", "input"),
        (32, "Equity", "input"),
        (33, "Balance Check", "formula", lambda c: f"=({c}20+{c}21+{c}22+{c}23+{c}24+{c}25+{c}26)-({c}27+{c}28+{c}29+{c}30+{c}31+{c}32)"),
        (35, "Cash Flow / Capital Allocation", "section"),
        (36, "Capex", "input"),
        (37, "Dividends", "input"),
        (38, "Buybacks", "input"),
        (39, "CFO (actual, optional)", "input"),
    ]

    for row_def in row_defs:
        row_num, label, kind, *formula = row_def
        ws[f"A{row_num}"] = label
        if kind == "section":
            for col in "ABCDE":
                style_cell(ws[f"{col}{row_num}"], fill=SECTION_FILL, bold=True)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
            style_cell(ws[f"A{row_num}"], fill=SECTION_FILL, bold=True)
            continue

        style_cell(ws[f"A{row_num}"], fill=OUTPUT_FILL if kind == "formula" else INPUT_FILL, bold=(kind == "formula"))
        for col in ["B", "C", "D"]:
            cell = ws[f"{col}{row_num}"]
            if kind == "input":
                style_cell(cell, fill=INPUT_FILL, num_format="#,##0.0")
            else:
                cell.value = formula[0](col)
                style_cell(cell, fill=FORMULA_FILL, num_format="#,##0.0")
        style_cell(ws[f"E{row_num}"], fill=FORMULA_FILL)

    comments = {
        6: "Top line / sales",
        7: "Use positive value",
        12: "Depreciation & amortization",
        16: "Use positive tax expense",
        33: "Should trend near zero",
        36: "Use positive cash outflow",
        39: "Optional actual CFO for reference",
    }
    for row_num, comment in comments.items():
        ws[f"E{row_num}"] = comment

    ws.freeze_panes = "B4"
    autofit_widths(ws, {"A": 28, "B": 14, "C": 14, "D": 14, "E": 24})


def add_whatif_sequences(ws):
    merge_title(ws, "What-If Sequence Map", "A1:H1")
    ws["A2"] = "Sequence -> driver inputs -> outputs. Use this sheet to understand what each stress sequence is trying to break."
    ws.merge_cells("A2:H2")
    style_cell(ws["A2"], fill=FORMULA_FILL)

    headers = ["Sequence", "Cause -> Effect Chain", "Driver Inputs In Model", "Outputs To Watch", "Light", "Base", "Severe", "Notes"]
    for idx, header in enumerate(headers, start=1):
        style_cell(ws.cell(row=3, column=idx, value=header), fill=SECTION_FILL, bold=True, align="center")

    row = 4
    for sequence in SEQUENCE_DEFINITIONS:
        ws[f"A{row}"] = sequence["name"]
        ws[f"B{row}"] = sequence["chain"]
        ws[f"C{row}"] = sequence["driver_inputs"]
        ws[f"D{row}"] = sequence["outputs"]
        ws[f"E{row}"] = summarize_severity(sequence["severities"]["Light"])
        ws[f"F{row}"] = summarize_severity(sequence["severities"]["Base"])
        ws[f"G{row}"] = summarize_severity(sequence["severities"]["Severe"])
        ws[f"H{row}"] = sequence["note"]
        for col in "ABCDEFGH":
            fill = OUTPUT_FILL if col == "A" else FORMULA_FILL
            style_cell(ws[f"{col}{row}"], fill=fill, wrap=True)
        ws.row_dimensions[row].height = 48
        row += 1

    ws["A26"] = "Key insight"
    style_cell(ws["A26"], fill=SECTION_FILL, bold=True)
    ws.merge_cells("B26:H26")
    ws["B26"] = "The most useful sequences are not always the extreme ones. The edge often comes from finding the combination of smaller changes that breaks the company."
    style_cell(ws["B26"], fill=FORMULA_FILL, wrap=True)
    ws.row_dimensions[26].height = 36

    ws.freeze_panes = "A4"
    autofit_widths(ws, {"A": 27, "B": 42, "C": 34, "D": 28, "E": 28, "F": 28, "G": 28, "H": 28})


def add_scenario_library(ws):
    merge_title(ws, "Scenario Library", "A1:V1")
    ws["A2"] = "Each row is a ready-to-use sequence and severity combination. Scenario_Setup pulls from this table."
    ws.merge_cells("A2:V2")
    style_cell(ws["A2"], fill=FORMULA_FILL)

    headers = [
        "Sequence",
        "Severity",
        "Scenario Key",
        "Cause -> Effect Chain",
        "Revenue Shock %",
        "Gross Margin Shock (bps)",
        "SG&A Mgmt Action %",
        "Other OpEx Shock %",
        "DSO Change",
        "DIO Change",
        "DPO Change",
        "Interest Shock (bps)",
        "ST Debt Rollover %",
        "LT Debt Due %",
        "LT Debt Rollover %",
        "Capex Cut %",
        "Dividend Cut %",
        "Buyback Cut %",
        "PP&E Impairment %",
        "Intangible Impairment %",
        "One-off Cash Charge % Rev",
        "Tax Rate %",
    ]
    for idx, header in enumerate(headers, start=1):
        style_cell(ws.cell(row=3, column=idx, value=header), fill=SECTION_FILL, bold=True, align="center")

    percent_cols = {5, 7, 8, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22}
    integer_cols = {6, 9, 10, 11, 12}

    row = 4
    for scenario in SCENARIO_ROWS:
        values = scenario["values"]
        output = [
            scenario["sequence"],
            scenario["severity"],
            scenario["key"],
            scenario["chain"],
            values["rev"],
            values["gm"],
            values["sga"],
            values["opex"],
            values["dso"],
            values["dio"],
            values["dpo"],
            values["rate"],
            values["st_roll"],
            values["lt_due"],
            values["lt_roll"],
            values["capex"],
            values["div"],
            values["buy"],
            values["ppe"],
            values["intg"],
            values["charge"],
            values["tax"],
        ]
        for col_idx, value in enumerate(output, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            style_cell(cell, fill=INPUT_FILL if col_idx >= 5 else FORMULA_FILL, wrap=(col_idx == 4))
            if col_idx in percent_cols:
                cell.number_format = "0.0%"
            elif col_idx in integer_cols:
                cell.number_format = "0"
        row += 1

    ws.freeze_panes = "A4"
    autofit_widths(
        ws,
        {
            "A": 27,
            "B": 10,
            "C": 38,
            "D": 48,
            "E": 16,
            "F": 18,
            "G": 18,
            "H": 18,
            "I": 12,
            "J": 12,
            "K": 12,
            "L": 18,
            "M": 16,
            "N": 14,
            "O": 16,
            "P": 14,
            "Q": 14,
            "R": 14,
            "S": 18,
            "T": 20,
            "U": 20,
            "V": 12,
        },
    )


def add_scenario_setup(ws):
    merge_title(ws, "Scenario Setup", "A1:E1")
    ws["A2"] = "Select a sequence and severity, then override any assumption in column D if you want a custom version."
    ws.merge_cells("A2:E2")
    style_cell(ws["A2"], fill=FORMULA_FILL)

    ws["A3"] = "Active Sequence"
    ws["B3"] = "Demand collapse (classic recession)"
    ws["A4"] = "Severity"
    ws["B4"] = "Base"
    ws["A5"] = "Scenario Key"
    ws["B5"] = '=B3&" | "&B4'
    ws["A6"] = "Description"
    ws["B6"] = f'=INDEX(Scenario_Library!$D$4:$D${SCENARIO_LAST_ROW},MATCH($B$5,Scenario_Library!$C$4:$C${SCENARIO_LAST_ROW},0))'

    for cell in ["A3", "A4", "A5", "A6"]:
        style_cell(ws[cell], fill=SECTION_FILL, bold=True)
    style_cell(ws["B3"], fill=ASSUMPTION_FILL, bold=True)
    style_cell(ws["B4"], fill=ASSUMPTION_FILL, bold=True)
    style_cell(ws["B5"], fill=FORMULA_FILL)
    style_cell(ws["B6"], fill=FORMULA_FILL, wrap=True)
    ws.merge_cells("B6:E6")
    ws.row_dimensions[6].height = 38

    sequence_dv = DataValidation(type="list", formula1=f"'WhatIf_Sequences'!$A$4:$A${WHATIF_LAST_ROW}", allow_blank=False)
    severity_dv = DataValidation(type="list", formula1='"Light,Base,Severe"', allow_blank=False)
    ws.add_data_validation(sequence_dv)
    ws.add_data_validation(severity_dv)
    sequence_dv.add(ws["B3"])
    severity_dv.add(ws["B4"])

    assumptions = [
        ("SG&A variable cost share", 0.40, "Portion of SG&A that flexes mechanically with revenue."),
        ("Fallback tax rate", '=IFERROR(Hist_Financials!D16/Hist_Financials!D15,0.25)', "Reference only. The active scenario tax rate drives the stress case."),
        ("Minimum cash buffer", 0.0, "Hard minimum cash level. Set this to covenant cash or your comfort floor."),
        ("Minimum EBIT / interest", 1.5, "Coverage threshold used for the distress flag."),
        ("Maximum net debt / EBITDA", 4.0, "Leverage threshold used for the distress flag."),
        ("Minimum current ratio", 1.0, "Liquidity threshold used for the distress flag."),
    ]

    ws["A8"] = "Company Settings"
    ws["B8"] = "Value"
    ws["C8"] = "Notes"
    style_cell(ws["A8"], fill=SECTION_FILL, bold=True)
    style_cell(ws["B8"], fill=SECTION_FILL, bold=True, align="center")
    style_cell(ws["C8"], fill=SECTION_FILL, bold=True)
    ws.merge_cells("C8:E8")

    for row, (label, value, note) in enumerate(assumptions, start=9):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"C{row}"] = note
        style_cell(ws[f"A{row}"], fill=OUTPUT_FILL)
        if row in (9, 10):
            num_format = "0.0%"
        elif row in (12, 13, 14):
            num_format = "0.00x"
        else:
            num_format = "#,##0.0"
        style_cell(ws[f"B{row}"], fill=ASSUMPTION_FILL, num_format=num_format)
        style_cell(ws[f"C{row}"], fill=FORMULA_FILL, wrap=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 30

    ws["A16"] = "Parameter"
    ws["B16"] = "Unit"
    ws["C16"] = "Library Value"
    ws["D16"] = "User Override"
    ws["E16"] = "Final Used"
    for cell in ["A16", "B16", "C16", "D16", "E16"]:
        style_cell(ws[cell], fill=SECTION_FILL, bold=True, align="center")

    for row, (label, unit, col_letter) in enumerate(PARAMETER_MAP, start=17):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = unit
        ws[f"C{row}"] = f'=INDEX(Scenario_Library!${col_letter}$4:${col_letter}${SCENARIO_LAST_ROW},MATCH($B$5,Scenario_Library!$C$4:$C${SCENARIO_LAST_ROW},0))'
        ws[f"D{row}"] = ""
        ws[f"E{row}"] = f'=IF(D{row}="",C{row},D{row})'
        style_cell(ws[f"A{row}"], fill=OUTPUT_FILL)
        style_cell(ws[f"B{row}"], fill=FORMULA_FILL, align="center")
        style_cell(ws[f"C{row}"], fill=FORMULA_FILL)
        style_cell(ws[f"D{row}"], fill=ASSUMPTION_FILL)
        style_cell(ws[f"E{row}"], fill=FORMULA_FILL)
        if unit == "%":
            for col in "CDE":
                ws[f"{col}{row}"].number_format = "0.0%"
        elif unit == "bps":
            for col in "CDE":
                ws[f"{col}{row}"].number_format = "0"
        else:
            for col in "CDE":
                ws[f"{col}{row}"].number_format = "0"

    ws.freeze_panes = "A16"
    autofit_widths(ws, {"A": 31, "B": 10, "C": 16, "D": 16, "E": 16})


def add_stress_model(ws):
    merge_title(ws, "Stress Model", "A1:D1")
    ws["A2"] = "Base case uses the latest historical year in Hist_Financials column D. Stress case applies the selected sequence and severity."
    ws.merge_cells("A2:D2")
    style_cell(ws["A2"], fill=FORMULA_FILL)

    for idx, header in enumerate(["Metric", "Base", "Stress", "Delta"], start=1):
        style_cell(ws.cell(row=4, column=idx, value=header), fill=SECTION_FILL, bold=True, align="center")

    sections = {5: "Income Statement", 22: "Working Capital", 31: "Cash Flow & Financing", 47: "Key Ratios"}
    for row_num, title in sections.items():
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        ws[f"A{row_num}"] = title
        style_cell(ws[f"A{row_num}"], fill=SECTION_FILL, bold=True)

    formulas = {
        6: ("Revenue", "=Hist_Financials!D6", "=B6*(1+Scenario_Setup!E17)"),
        7: ("Gross margin %", "=IFERROR((Hist_Financials!D6-Hist_Financials!D7)/Hist_Financials!D6,0)", "=MAX(-50%,MIN(90%,B7+Scenario_Setup!E18/10000))"),
        8: ("Gross profit", "=Hist_Financials!D8", "=C6*C7"),
        9: ("SG&A", "=Hist_Financials!D9", "=MAX(0,B9*(1+Scenario_Setup!$B$9*Scenario_Setup!E17+Scenario_Setup!E19))"),
        10: ("Other OpEx", "=Hist_Financials!D10", "=MAX(0,B10*(1+Scenario_Setup!E20))"),
        11: ("EBITDA", "=Hist_Financials!D11", "=C8-C9-C10"),
        12: ("D&A", "=Hist_Financials!D12", "=B12"),
        13: ("EBIT", "=Hist_Financials!D13", "=C11-C12"),
        14: ("Interest rate %", "=IFERROR(Hist_Financials!D14/(Hist_Financials!D27+Hist_Financials!D30),0)", "=MAX(0,B14+Scenario_Setup!E24/10000)"),
        15: ("Interest expense", "=Hist_Financials!D14", "=C14*((B43+C43)/2)"),
        16: ("PP&E impairment", "=0", "=Hist_Financials!D24*Scenario_Setup!E31"),
        17: ("Intangible impairment", "=0", "=Hist_Financials!D25*Scenario_Setup!E32"),
        18: ("One-off cash charge", "=0", "=Hist_Financials!D6*Scenario_Setup!E33"),
        19: ("Pretax income", "=Hist_Financials!D15", "=C13-C15-C16-C17-C18"),
        20: ("Taxes", "=Hist_Financials!D16", "=MAX(0,C19)*Scenario_Setup!E34"),
        21: ("Net income", "=Hist_Financials!D17", "=C19-C20"),
        23: ("DSO", "=IFERROR(Hist_Financials!D21/Hist_Financials!D6*365,0)", "=MAX(0,B23+Scenario_Setup!E21)"),
        24: ("DIO", "=IFERROR(Hist_Financials!D22/Hist_Financials!D7*365,0)", "=MAX(0,B24+Scenario_Setup!E22)"),
        25: ("DPO", "=IFERROR(Hist_Financials!D28/Hist_Financials!D7*365,0)", "=MAX(0,B25+Scenario_Setup!E23)"),
        26: ("Accounts receivable", "=Hist_Financials!D21", "=C6*C23/365"),
        27: ("Inventory", "=Hist_Financials!D22", "=(C6-C8)*C24/365"),
        28: ("Accounts payable", "=Hist_Financials!D28", "=(C6-C8)*C25/365"),
        29: ("Other current assets", "=Hist_Financials!D23", "=B29"),
        30: ("Other current liabilities", "=Hist_Financials!D29", "=B30"),
        32: ("CFO (modeled)", '=IF(Hist_Financials!D39<>"",Hist_Financials!D39,Hist_Financials!D17+Hist_Financials!D12)', "=C21+C12+C16+C17-(C26-B26)-(C27-B27)+(C28-B28)"),
        33: ("Capex", "=Hist_Financials!D36", "=B33*(1-Scenario_Setup!E28)"),
        34: ("Dividends", "=Hist_Financials!D37", "=B34*(1-Scenario_Setup!E29)"),
        35: ("Buybacks", "=Hist_Financials!D38", "=B35*(1-Scenario_Setup!E30)"),
        36: ("Cash before financing", "=Hist_Financials!D20", "=Hist_Financials!D20+C32-C33-C34-C35"),
        37: ("Mandatory ST debt repayment", "=0", "=Hist_Financials!D27*(1-Scenario_Setup!E25)"),
        38: ("Mandatory LT debt repayment", "=0", "=Hist_Financials!D30*Scenario_Setup!E26*(1-Scenario_Setup!E27)"),
        39: ("Ending cash", "=Hist_Financials!D20", "=C36-C37-C38"),
        40: ("Funding gap to minimum cash", "=0", "=MAX(0,Scenario_Setup!$B$11-C39)"),
        41: ("Ending ST debt", "=Hist_Financials!D27", "=Hist_Financials!D27-C37"),
        42: ("Ending LT debt", "=Hist_Financials!D30", "=Hist_Financials!D30-C38"),
        43: ("Total ending debt", "=B41+B42", "=C41+C42"),
        44: ("Ending equity", "=Hist_Financials!D32", "=Hist_Financials!D32+C21-C34-C35"),
        45: ("Ending PP&E", "=Hist_Financials!D24", "=MAX(0,Hist_Financials!D24+C33-C12-C16)"),
        46: ("Ending intangibles", "=Hist_Financials!D25", "=MAX(0,Hist_Financials!D25-C17)"),
        48: ("Net debt / EBITDA", '=IFERROR((B43-Hist_Financials!D20)/MAX(B11,0.01),0)', '=IFERROR((C43-C39+C40)/MAX(C11,0.01),0)'),
        49: ("EBIT / interest", '=IFERROR(B13/MAX(B15,0.01),999)', '=IFERROR(C13/MAX(C15,0.01),999)'),
        50: ("Current ratio", '=IFERROR((Hist_Financials!D20+Hist_Financials!D21+Hist_Financials!D22+Hist_Financials!D23)/(Hist_Financials!D28+Hist_Financials!D29+Hist_Financials!D27),0)', '=IFERROR((MAX(0,C39)+C26+C27+C29)/(C28+C30+MAX(0,C41)),0)'),
    }

    percent_rows = {7, 14}
    ratio_rows = {48, 49, 50}
    day_rows = {23, 24, 25}

    for row_num, (label, base_formula, stress_formula) in formulas.items():
        ws[f"A{row_num}"] = label
        ws[f"B{row_num}"] = base_formula
        ws[f"C{row_num}"] = stress_formula
        ws[f"D{row_num}"] = f"=C{row_num}-B{row_num}"
        style_cell(ws[f"A{row_num}"], fill=OUTPUT_FILL)
        style_cell(ws[f"B{row_num}"], fill=FORMULA_FILL)
        style_cell(ws[f"C{row_num}"], fill=STRESS_FILL)
        style_cell(ws[f"D{row_num}"], fill=FORMULA_FILL)
        if row_num in percent_rows:
            for col in "BCD":
                ws[f"{col}{row_num}"].number_format = "0.0%"
        elif row_num in ratio_rows:
            for col in "BCD":
                ws[f"{col}{row_num}"].number_format = "0.00x"
        elif row_num in day_rows:
            for col in "BCD":
                ws[f"{col}{row_num}"].number_format = "0"
        else:
            for col in "BCD":
                ws[f"{col}{row_num}"].number_format = "#,##0.0"

    ws.freeze_panes = "B5"
    autofit_widths(ws, {"A": 31, "B": 15, "C": 15, "D": 15})


def add_dashboard(ws):
    merge_title(ws, "Dashboard", "A1:G1")
    ws["A3"] = "Sequence"
    ws["B3"] = "=Scenario_Setup!B3"
    ws["A4"] = "Severity"
    ws["B4"] = "=Scenario_Setup!B4"
    ws["A5"] = "Description"
    ws["B5"] = "=Scenario_Setup!B6"

    for cell in ["A3", "A4", "A5"]:
        style_cell(ws[cell], fill=SECTION_FILL, bold=True)
    style_cell(ws["B3"], fill=OUTPUT_FILL, bold=True)
    style_cell(ws["B4"], fill=OUTPUT_FILL, bold=True)
    style_cell(ws["B5"], fill=FORMULA_FILL, wrap=True)
    ws.merge_cells("B5:G5")
    ws.row_dimensions[5].height = 38

    headers = ["Metric", "Base", "Stress", "Delta", "Threshold", "Status"]
    for idx, header in enumerate(headers, start=1):
        style_cell(ws.cell(row=7, column=idx, value=header), fill=SECTION_FILL, bold=True, align="center")

    metrics = [
        ("Revenue", "Stress_Model!B6", "Stress_Model!C6", "n", ""),
        ("EBITDA", "Stress_Model!B11", "Stress_Model!C11", "n", ""),
        ("Net income", "Stress_Model!B21", "Stress_Model!C21", "n", ""),
        ("CFO (modeled)", "Stress_Model!B32", "Stress_Model!C32", "n", ""),
        ("Ending cash", "Stress_Model!B39", "Stress_Model!C39", "n", ">= Scenario_Setup!B11"),
        ("Funding gap", "Stress_Model!B40", "Stress_Model!C40", "n", "= 0"),
        ("Net debt / EBITDA", "Stress_Model!B48", "Stress_Model!C48", "r", "<= Scenario_Setup!B13"),
        ("EBIT / interest", "Stress_Model!B49", "Stress_Model!C49", "r", ">= Scenario_Setup!B12"),
        ("Current ratio", "Stress_Model!B50", "Stress_Model!C50", "r", ">= Scenario_Setup!B14"),
        ("Ending equity", "Stress_Model!B44", "Stress_Model!C44", "n", "> 0"),
    ]

    start_row = 8
    for row, (label, base_ref, stress_ref, metric_type, threshold) in enumerate(metrics, start=start_row):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = f"={base_ref}"
        ws[f"C{row}"] = f"={stress_ref}"
        ws[f"D{row}"] = f"=C{row}-B{row}"
        ws[f"E{row}"] = threshold
        if label == "Ending cash":
            ws[f"F{row}"] = '=IF(C12>=Scenario_Setup!B11,"OK","FAIL")'
        elif label == "Funding gap":
            ws[f"F{row}"] = '=IF(C13=0,"OK","FAIL")'
        elif label == "Net debt / EBITDA":
            ws[f"F{row}"] = '=IF(C14<=Scenario_Setup!B13,"OK","FAIL")'
        elif label == "EBIT / interest":
            ws[f"F{row}"] = '=IF(C15>=Scenario_Setup!B12,"OK","FAIL")'
        elif label == "Current ratio":
            ws[f"F{row}"] = '=IF(C16>=Scenario_Setup!B14,"OK","FAIL")'
        elif label == "Ending equity":
            ws[f"F{row}"] = '=IF(C17>0,"OK","FAIL")'
        else:
            ws[f"F{row}"] = '"INFO"'
        style_cell(ws[f"A{row}"], fill=OUTPUT_FILL)
        style_cell(ws[f"B{row}"], fill=FORMULA_FILL)
        style_cell(ws[f"C{row}"], fill=STRESS_FILL)
        style_cell(ws[f"D{row}"], fill=FORMULA_FILL)
        style_cell(ws[f"E{row}"], fill=FORMULA_FILL)
        style_cell(ws[f"F{row}"], fill=FORMULA_FILL, bold=True, align="center")
        if metric_type == "r":
            for col in "BCD":
                ws[f"{col}{row}"].number_format = "0.00x"
        else:
            for col in "BCD":
                ws[f"{col}{row}"].number_format = "#,##0.0"

    ws["A20"] = "Overall assessment"
    ws["B20"] = '=IF(OR(F13="FAIL",F14="FAIL",F17="FAIL"),"High distress",IF(OR(F15="FAIL",F16="FAIL"),"Elevated risk","Within selected thresholds"))'
    style_cell(ws["A20"], fill=SECTION_FILL, bold=True)
    style_cell(ws["B20"], fill=OUTPUT_FILL, bold=True)
    ws.merge_cells("B20:D20")

    ws["A22"] = "Active driver inputs"
    style_cell(ws["A22"], fill=SECTION_FILL, bold=True)
    assumption_refs = [
        ("Revenue shock %", "Scenario_Setup!E17", "%"),
        ("Gross margin shock (bps)", "Scenario_Setup!E18", "bps"),
        ("DSO change", "Scenario_Setup!E21", "days"),
        ("DIO change", "Scenario_Setup!E22", "days"),
        ("DPO change", "Scenario_Setup!E23", "days"),
        ("Interest shock (bps)", "Scenario_Setup!E24", "bps"),
        ("ST debt rollover %", "Scenario_Setup!E25", "%"),
        ("LT debt due %", "Scenario_Setup!E26", "%"),
        ("LT debt rollover %", "Scenario_Setup!E27", "%"),
        ("Capex cut %", "Scenario_Setup!E28", "%"),
        ("One-off cash charge % rev", "Scenario_Setup!E33", "%"),
    ]
    for row, (label, ref, unit) in enumerate(assumption_refs, start=23):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = f"={ref}"
        style_cell(ws[f"A{row}"], fill=OUTPUT_FILL)
        style_cell(ws[f"B{row}"], fill=FORMULA_FILL)
        if unit == "%":
            ws[f"B{row}"].number_format = "0.0%"
        else:
            ws[f"B{row}"].number_format = "0"

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Base vs Stress"
    chart.y_axis.title = "Metric"
    chart.x_axis.title = "Value"
    data = Reference(ws, min_col=2, max_col=3, min_row=8, max_row=12)
    cats = Reference(ws, min_col=1, min_row=8, max_row=12)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 10
    chart.legend.position = "r"
    ws.add_chart(chart, "D22")

    ws.conditional_formatting.add("F8:F17", CellIsRule(operator="equal", formula=['"OK"'], fill=PASS_FILL))
    ws.conditional_formatting.add("F8:F17", CellIsRule(operator="equal", formula=['"FAIL"'], fill=FAIL_FILL))
    ws.conditional_formatting.add("F8:F17", CellIsRule(operator="equal", formula=['"INFO"'], fill=INFO_FILL))

    ws.freeze_panes = "A7"
    autofit_widths(ws, {"A": 25, "B": 18, "C": 18, "D": 16, "E": 18, "F": 12, "G": 14})


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

    readme = wb.create_sheet("README")
    hist = wb.create_sheet("Hist_Financials")
    whatif = wb.create_sheet("WhatIf_Sequences")
    library = wb.create_sheet("Scenario_Library")
    setup = wb.create_sheet("Scenario_Setup")
    model = wb.create_sheet("Stress_Model")
    dashboard = wb.create_sheet("Dashboard")

    add_readme(readme)
    add_hist_financials(hist)
    add_whatif_sequences(whatif)
    add_scenario_library(library)
    add_scenario_setup(setup)
    add_stress_model(model)
    add_dashboard(dashboard)

    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    build_workbook()
